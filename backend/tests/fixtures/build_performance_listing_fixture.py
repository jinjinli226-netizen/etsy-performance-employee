from __future__ import annotations

from datetime import datetime, timezone
from io import BytesIO
from pathlib import Path
from zipfile import ZIP_DEFLATED, ZipFile, ZipInfo

from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, PatternFill


OUTPUT = Path(__file__).with_name("performance-listing-template.xlsx")
PNG = bytes.fromhex(
    "89504e470d0a1a0a0000000d49484452000000010000000108060000001f15c489"
    "0000000d49444154789c6360606060000000050001a5f645400000000049454e44ae426082"
)
FIXED_ZIP_TIME = (2024, 1, 1, 0, 0, 0)


def canonicalize_xlsx(path: Path) -> None:
    with ZipFile(path, "r") as source:
        entries = [(item.filename, source.read(item.filename)) for item in source.infolist()]
    buffer = BytesIO()
    with ZipFile(buffer, "w", compression=ZIP_DEFLATED, compresslevel=9) as destination:
        for name, content in sorted(entries):
            info = ZipInfo(name, date_time=FIXED_ZIP_TIME)
            info.compress_type = ZIP_DEFLATED
            info.external_attr = 0o600 << 16
            destination.writestr(info, content)
    path.write_bytes(buffer.getvalue())


def build(path: Path = OUTPUT) -> None:
    workbook = Workbook()
    fixed_time = datetime(2024, 1, 1, tzinfo=timezone.utc)
    workbook.properties.created = fixed_time
    workbook.properties.modified = fixed_time
    workbook.properties.creator = "Synthetic test fixture"
    workbook.properties.lastModifiedBy = "Synthetic test fixture"
    sheet = workbook.active
    sheet.title = "Products"
    headers = (
        "SKU",
        "Product notes",
        "Cost price",
        "Logistics status",
        "head titles",
        "13 tags",
        "SPECIFICATION",
        "Category",
        "Instructions for buyers",
    )
    for column, header in enumerate(headers, 1):
        cell = sheet.cell(3, column, header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")
    sheet.append([])
    sheet.cell(4, 1, "instruction")
    sheet.cell(4, 2, "Please fill one product per row")
    sheet.cell(5, 1, "SYN-001")
    sheet.cell(5, 2, "Blue template-inspired dance costume")
    sheet.cell(5, 3, 10)
    sheet.cell(5, 4, "internal-air")
    sheet.cell(6, 1, "SYN-002")
    sheet.cell(6, 2, "Red sequin stage costume")
    sheet.cell(6, 3, 20)
    sheet.cell(6, 4, "internal-sea")
    sheet["B5"].hyperlink = "https://example.invalid/synthetic-product"
    sheet["B5"].fill = PatternFill("solid", fgColor="00FF00")
    sheet["D10"] = "=1+1"
    sheet.column_dimensions["B"].width = 42
    sheet.row_dimensions[5].height = 30
    for row in (5, 6):
        image = Image(BytesIO(PNG))
        image.width = image.height = 1
        sheet.add_image(image, f"B{row}")
    workbook.create_sheet("Reference")["A1"] = "Synthetic fixture only"
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)
    canonicalize_xlsx(path)


if __name__ == "__main__":
    build()
