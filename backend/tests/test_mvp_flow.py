from __future__ import annotations

import hashlib
import shutil
from datetime import UTC, datetime
from pathlib import Path

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy import select

from app.core.config import Settings
from app.db.init_db import init_db
from app.db.models import Conversation, ImportedEvidenceFingerprint, KnowledgeCandidate, KnowledgePattern, RuleVersion
from app.db.session import create_engine_for_url, create_session_factory
from app.knowledge.schemas import CandidateInput, EvidenceReference, KnowledgeStatus
from app.knowledge.service import KnowledgeService, KnowledgeValidationError
from app.main import create_app
from app.migration.importer import MigrationImporter
from tests.fakes.mvp_runtime import EmployeeSkillRunner, TeachingHermes


ROOT = Path(__file__).resolve().parents[2]
FIXTURE = Path(__file__).parent / "fixtures" / "performance-listing-template.xlsx"
OUTPUT_HEADERS = ("head titles", "13 tags", "SPECIFICATION", "Category", "Instructions for buyers")


def sha256(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def wait_for_chat(client: TestClient, operation_id: str) -> None:
    with client.stream("GET", f"/api/events/{operation_id}") as response:
        assert response.status_code == 200
        assert any('"status":"completed"' in line for line in response.iter_lines() if line.startswith("data:"))


def wait_for_job(client: TestClient, job_id: str) -> dict:
    import time

    for _ in range(300):
        response = client.get(f"/api/excel-jobs/{job_id}")
        assert response.status_code == 200
        job = response.json()
        if job["status"] in {"completed", "failed", "cancelled"}:
            return job
        time.sleep(0.01)
    raise AssertionError("Excel job did not finish")


def test_complete_local_mvp_flow_uses_employee_skill_and_migrates_portably(tmp_path: Path) -> None:
    source_hash = sha256(FIXTURE)
    settings = Settings(data_dir=tmp_path / "source-data", database_url=f"sqlite:///{tmp_path / 'source.db'}")
    app = create_app(settings=settings, employee=TeachingHermes(), excel_runner=EmployeeSkillRunner())

    with TestClient(app) as client:
        conversation = client.post("/api/conversations", json={"title": "竞品教学"})
        assert conversation.status_code == 201
        conversation_id = conversation.json()["id"]
        sent = client.post(
            f"/api/conversations/{conversation_id}/messages",
            json={"content": "请学习 https://www.etsy.com/listing/700001/example", "learning_mode": True},
        )
        assert sent.status_code == 202
        wait_for_chat(client, sent.json()["operation_id"])

        candidates = client.get("/api/knowledge/candidates").json()
        assert candidates["total"] == 1 and candidates["items"][0]["status"] == "proposed"
        approved = client.post(f"/api/knowledge/candidates/{candidates['items'][0]['id']}/approve")
        assert approved.status_code == 200 and approved.json()["status"] == "active"
        pattern_id = approved.json()["id"]

        with app.state.session_factory() as session:
            first = session.get(KnowledgeCandidate, candidates["items"][0]["id"])
            evidence_id = first.evidence_ids[0]
            source_timestamp = first.source_timestamps[evidence_id]
        second = app.state.knowledge_service.ingest_candidate(
            CandidateInput(
                kind="title_structure",
                abstract="Place buyer intent before costume style and audience wording.",
                confidence=0.92,
                evidence_refs=[EvidenceReference(evidence_id=evidence_id, source_timestamp=datetime.fromisoformat(source_timestamp))],
            ),
            actor="owner",
            trace_id="mvp-second-rule",
        )
        second_approved = client.post(f"/api/knowledge/candidates/{second.id}/approve")
        assert second_approved.status_code == 200

        uploaded = client.post(
            "/api/excel-jobs",
            files={"file": ("products.xlsx", FIXTURE.read_bytes(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
        assert uploaded.status_code == 202
        job = wait_for_job(client, uploaded.json()["id"])
        assert job["status"] == "completed", job
        assert job["source_sha256"] == source_hash == sha256(FIXTURE)
        downloaded = client.get(f"/api/excel-jobs/{job['id']}/download")
        assert downloaded.status_code == 200
        artifact = tmp_path / "downloaded.xlsx"
        artifact.write_bytes(downloaded.content)

        workbook = load_workbook(artifact, data_only=False)
        try:
            sheet = workbook["Products"]
            assert workbook.sheetnames == ["Products", "Reference"]
            assert sheet["D10"].value == "=1+1"
            assert sheet["B5"].hyperlink.target == "https://example.invalid/synthetic-product"
            assert len(sheet._images) == 2
            for row_number, sku in ((5, "SYN-001"), (6, "SYN-002")):
                values = [sheet.cell(row_number, column).value for column in range(5, 10)]
                assert all(isinstance(value, str) and value for value in values)
                assert sku in values[2]
                assert len(values[1].split(", ")) == 13
            assert "SYN-002" not in sheet["G5"].value and "SYN-001" not in sheet["G6"].value
        finally:
            workbook.close()

        copied = tmp_path / "near-copy.xlsx"
        shutil.copyfile(artifact, copied)
        workbook = load_workbook(copied)
        workbook["Products"]["E5"] = "Protected competitor crystal fringe wording for a theatrical stage garment"
        workbook.save(copied)
        workbook.close()
        with pytest.raises(KnowledgeValidationError):
            app.state.knowledge_service.validate_generated_workbook(copied)

        package = tmp_path / "employee-package.zip"
        exported = app.state.migration_exporter.export(package, created_at=datetime(2026, 8, 14, tzinfo=UTC))
        assert exported.path == package and exported.record_counts["conversations"] == 1

    target_engine = create_engine_for_url(f"sqlite:///{tmp_path / 'target.db'}")
    init_db(target_engine)
    target_factory = create_session_factory(target_engine)
    importer = MigrationImporter(
        target_factory,
        repository_assets=ROOT / "employee",
        workspace=tmp_path / "target-import-workspace",
    )
    dry_run = importer.import_package(package, dry_run=True)
    assert dry_run.dry_run is True
    imported = importer.import_package(package)
    assert imported.imported is True and imported.credential_status == "pending"

    target_knowledge = KnowledgeService(target_factory, export_dir=tmp_path / "target-trust")
    with target_factory() as session:
        assert session.scalar(select(Conversation)) is not None
        restored_pattern = session.scalar(select(KnowledgePattern))
        assert restored_pattern is not None
        assert len(list(session.scalars(select(RuleVersion).where(RuleVersion.pattern_id == restored_pattern.id)))) == 2
        assert session.scalar(select(ImportedEvidenceFingerprint).where(ImportedEvidenceFingerprint.public_id == evidence_id)) is not None
        active_before = session.scalar(select(RuleVersion).where(RuleVersion.pattern_id == restored_pattern.id, RuleVersion.status == KnowledgeStatus.ACTIVE))
        assert active_before is not None
    rollback = target_knowledge.rollback_pattern(
        restored_pattern.id,
        actor="owner",
        expected_rule_version=active_before.version,
    )
    assert rollback.status is KnowledgeStatus.ACTIVE and rollback.rule_version != active_before.version
    target_engine.dispose()
