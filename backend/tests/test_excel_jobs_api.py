from __future__ import annotations

import asyncio
import hashlib
import json
import os
import shutil
import stat
import sys
import time
from pathlib import Path
from uuid import NAMESPACE_URL, uuid5

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy import select, text

from app.core.config import Settings
from app.excel_jobs.runner import (
    RunnerRequest,
    SubprocessExcelRunner,
    WorkerProtocolError,
    WorkerResult,
    build_employee_command,
)
from app.excel_jobs.storage import StorageError, publish_artifact
from app.excel_jobs import storage as excel_storage
from app.main import create_app
from app.db.init_db import init_db
from app.db.models import Artifact, ExcelJob, JobEvent
from app.db.session import create_engine_for_url, create_session_factory
from app.excel_jobs.schemas import JobStatus


FIXTURE = Path(__file__).parent / "fixtures" / "performance-listing-template.xlsx"


class FakeExcelRunner:
    def __init__(self, *, fail: bool = False, malformed: bool = False, block: bool = False) -> None:
        self.fail = fail
        self.malformed = malformed
        self.block = block
        self.calls: list[RunnerRequest] = []
        self.cancelled: set[str] = set()

    async def run(self, request: RunnerRequest, emit) -> WorkerResult:
        self.calls.append(request)
        await emit({"event": "started"})
        if self.malformed:
            await emit({"not_event": "bad"})
        if self.block:
            while request.public_id not in self.cancelled:
                await asyncio.sleep(0.01)
            raise asyncio.CancelledError
        output = request.operation_dir / "generated.xlsx"
        shutil.copyfile(request.source_path, output)
        if self.fail:
            raise RuntimeError("private stderr secret")
        return WorkerResult(output_path=output, output_sha256=sha256(output))

    async def cancel(self, public_id: str) -> None:
        self.cancelled.add(public_id)

    async def shutdown(self) -> None:
        return None


class RealProtocolRunner(FakeExcelRunner):
    async def run(self, request: RunnerRequest, emit) -> WorkerResult:
        self.calls.append(request)
        await emit({"event": "started"})
        await emit({"event": "row_started", "row_id": "row-1", "row_number": 3})
        await emit({"event": "row_completed", "row_id": "row-1", "row_number": 3})
        await emit({"event": "row_started", "row_id": "row-2", "row_number": 4})
        await emit({"event": "row_completed", "row_id": "row-2", "row_number": 4})
        output = request.operation_dir / "generated.xlsx"
        shutil.copyfile(request.source_path, output)
        result = WorkerResult(output, sha256(output))
        await emit({"event": "completed", "output_path": str(output), "output_sha256": result.output_sha256})
        return result


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for chunk in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


@pytest.fixture
def api(tmp_path):
    runner = FakeExcelRunner()
    settings = Settings(
        data_dir=tmp_path / "data",
        database_url=f"sqlite:///{tmp_path / 'api.db'}",
        max_excel_upload_bytes=5 * 1024 * 1024,
    )
    app = create_app(settings=settings, excel_runner=runner)
    with TestClient(app) as client:
        yield client, runner, settings, app


def upload(client: TestClient, name: str = "products.xlsx"):
    return client.post(
        "/api/excel-jobs",
        files={
            "file": (
                name,
                FIXTURE.read_bytes(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )


def wait_terminal(client: TestClient, public_id: str) -> dict:
    for _ in range(200):
        payload = client.get(f"/api/excel-jobs/{public_id}").json()
        if payload["status"] in {"completed", "failed", "cancelled"}:
            return payload
        import time

        time.sleep(0.01)
    raise AssertionError("job did not reach a terminal state")


def test_upload_runs_employee_job_persists_events_and_downloads_new_artifact(api) -> None:
    client, runner, settings, _ = api
    source_digest = sha256(FIXTURE)

    created = upload(client, "../../unsafe name.xlsx")
    assert created.status_code == 202
    job = created.json()
    assert job["status"] in {"queued", "running", "completed"}
    assert len(job["id"]) == 36
    assert "path" not in json.dumps(job).casefold()

    terminal = wait_terminal(client, job["id"])
    assert terminal["status"] == "completed"
    assert terminal["source_sha256"] == source_digest
    assert terminal["progress_percent"] == 100
    assert terminal["artifact"]["sha256"]

    listing = client.get("/api/excel-jobs", params={"limit": 10, "offset": 0})
    assert listing.status_code == 200
    assert listing.json()["total"] == 1
    assert listing.json()["items"][0]["id"] == job["id"]

    downloaded = client.get(f"/api/excel-jobs/{job['id']}/download")
    assert downloaded.status_code == 200
    assert downloaded.content != b""
    assert "attachment" in downloaded.headers["content-disposition"].casefold()
    assert ".." not in downloaded.headers["content-disposition"]

    call = runner.calls[0]
    assert call.source_path.parent.name == "source"
    assert call.source_path.name == "source.xlsx"
    assert sha256(call.source_path) == source_digest
    assert call.source_path.read_bytes() == FIXTURE.read_bytes()
    assert not call.source_path.stat().st_mode & stat.S_IWRITE
    assert call.source_path.is_relative_to(settings.data_dir / "excel-jobs")
    assert call.knowledge_path.is_relative_to(settings.data_dir)


@pytest.mark.parametrize(
    ("filename", "content", "expected"),
    [
        ("products.xlsm", b"PK bad", 415),
        ("products.xlsx", b"not an OOXML package", 422),
        ("products.xlsx.exe", b"PK bad", 415),
    ],
)
def test_upload_rejects_unsupported_or_disguised_files(api, filename, content, expected) -> None:
    client, _, _, _ = api
    response = client.post("/api/excel-jobs", files={"file": (filename, content)})
    assert response.status_code == expected
    assert client.get("/api/excel-jobs").json()["total"] == 0


def test_upload_returns_507_before_queue_or_worker_when_knowledge_capacity_exceeded(api, monkeypatch) -> None:
    client, runner, _, app = api
    from app.knowledge.service import KnowledgeCapacityError

    def exceeded():
        raise KnowledgeCapacityError("private capacity details")

    monkeypatch.setattr(app.state.knowledge_service, "require_capacity_ready", exceeded)
    response = upload(client)
    assert response.status_code == 507
    assert response.json()["detail"]["code"] == "knowledge_capacity_exceeded"
    assert runner.calls == []
    assert client.get("/api/excel-jobs").json()["total"] == 0


def test_queued_job_fails_with_capacity_code_before_worker_starts(api, monkeypatch) -> None:
    client, runner, _, app = api
    from app.knowledge.service import KnowledgeCapacityError
    from app.excel_jobs.storage import store_upload
    from uuid import uuid4

    public_id = uuid4()

    async def prepare():
        class Upload:
            filename = "products.xlsx"
            content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

            async def read(self, size=-1):
                if getattr(self, "done", False):
                    return b""
                self.done = True
                return FIXTURE.read_bytes()

            async def close(self):
                return None

        return await store_upload(Upload(), root=app.state.settings.data_dir / "excel-jobs", public_id=public_id, max_bytes=5 * 1024 * 1024)

    stored = asyncio.run(prepare())
    job = app.state.excel_job_service.create_job(stored, "products.xlsx")

    def exceeded():
        raise KnowledgeCapacityError("private capacity details")

    monkeypatch.setattr(app.state.knowledge_service, "require_capacity_ready", exceeded)
    asyncio.run(app.state.excel_job_service._execute(job.id))
    terminal = client.get(f"/api/excel-jobs/{job.id}").json()
    assert terminal["status"] == "failed"
    assert terminal["error"]["code"] == "knowledge_capacity_exceeded"
    assert runner.calls == []


def test_upload_streaming_hard_cap_returns_413_and_leaves_no_job(tmp_path) -> None:
    runner = FakeExcelRunner()
    settings = Settings(
        data_dir=tmp_path / "data",
        database_url=f"sqlite:///{tmp_path / 'api.db'}",
        max_excel_upload_bytes=64,
    )
    with TestClient(create_app(settings=settings, excel_runner=runner)) as client:
        response = upload(client)
        assert response.status_code == 413
        assert client.get("/api/excel-jobs").json()["total"] == 0
        assert not list((settings.data_dir / "excel-jobs").glob("*"))


def test_failed_worker_hides_internal_error_and_never_exposes_partial(tmp_path) -> None:
    runner = FakeExcelRunner(fail=True)
    settings = Settings(data_dir=tmp_path / "data", database_url=f"sqlite:///{tmp_path / 'api.db'}")
    app = create_app(settings=settings, excel_runner=runner)
    with TestClient(app) as client:
        job = upload(client).json()
        terminal = wait_terminal(client, job["id"])
        assert terminal["status"] == "failed"
        assert terminal["error"]["code"] == "worker_failed"
        assert "secret" not in json.dumps(terminal).casefold()
        assert client.get(f"/api/excel-jobs/{job['id']}/download").status_code == 409
        workspace = settings.data_dir / "excel-jobs" / job["id"]
        assert not list((workspace / "artifacts").glob("*.xlsx"))
        assert (workspace / "source" / "source.xlsx").exists()


def test_malformed_progress_event_fails_job_safely(tmp_path) -> None:
    runner = FakeExcelRunner(malformed=True)
    settings = Settings(data_dir=tmp_path / "data", database_url=f"sqlite:///{tmp_path / 'api.db'}")
    with TestClient(create_app(settings=settings, excel_runner=runner)) as client:
        job = upload(client).json()
        terminal = wait_terminal(client, job["id"])
        assert terminal["status"] == "failed"
        assert terminal["error"]["code"] == "invalid_worker_event"


def test_real_multi_row_worker_protocol_reaches_terminal_and_persists_events(tmp_path) -> None:
    settings = Settings(data_dir=tmp_path / "data", database_url=f"sqlite:///{tmp_path / 'api.db'}")
    with TestClient(create_app(settings=settings, excel_runner=RealProtocolRunner())) as client:
        job = upload(client).json()
        terminal = wait_terminal(client, job["id"])
        assert terminal["status"] == "completed"
        with client.stream("GET", f"/api/excel-jobs/{job['id']}/events") as response:
            event_types = [line[7:] for line in response.iter_lines() if line.startswith("event: ")]
        assert event_types == [
            "queued", "running", "worker_started", "worker_row_started",
            "worker_row_completed", "worker_row_started", "worker_row_completed",
            "worker_completed", "completed",
        ]


def test_event_persist_failure_is_caught_and_job_becomes_failed(tmp_path, monkeypatch) -> None:
    settings = Settings(data_dir=tmp_path / "data", database_url=f"sqlite:///{tmp_path / 'api.db'}")
    app = create_app(settings=settings, excel_runner=RealProtocolRunner())
    with TestClient(app) as client:
        original = app.state.excel_job_service._persist_worker_event_sync
        calls = 0

        def fail_once(public_id, kind, event):
            nonlocal calls
            calls += 1
            if calls == 2:
                raise OSError("database write failed")
            return original(public_id, kind, event)

        monkeypatch.setattr(app.state.excel_job_service, "_persist_worker_event_sync", fail_once)
        job = upload(client).json()
        terminal = wait_terminal(client, job["id"])
        assert terminal["status"] == "failed"
        assert terminal["error"]["code"] == "worker_failed"


def test_sse_replays_persisted_events_after_last_event_id(api) -> None:
    client, _, _, _ = api
    job = upload(client).json()
    wait_terminal(client, job["id"])

    with client.stream("GET", f"/api/excel-jobs/{job['id']}/events") as response:
        assert response.status_code == 200
        lines = list(response.iter_lines())
    ids = [int(line[4:]) for line in lines if line.startswith("id: ")]
    assert ids == sorted(ids)
    assert len(ids) >= 3

    with client.stream(
        "GET",
        f"/api/excel-jobs/{job['id']}/events",
        headers={"Last-Event-ID": str(ids[-2])},
    ) as response:
        replay = list(response.iter_lines())
    assert [int(line[4:]) for line in replay if line.startswith("id: ")] == [ids[-1]]


def test_sse_replays_more_than_two_pages_before_terminal_event(tmp_path) -> None:
    database = tmp_path / "many-events.db"
    engine = create_engine_for_url(f"sqlite:///{database}")
    init_db(engine)
    factory = create_session_factory(engine)
    public_id = "00000000-0000-4000-8000-000000000099"
    with factory() as session:
        job = ExcelJob(
            public_id=public_id, source_filename="events.xlsx", source_sha256="a" * 64,
            source_size_bytes=1, status=JobStatus.COMPLETED, progress_percent=100,
        )
        job.events.extend(JobEvent(event_type="progress", payload={"sequence": index}) for index in range(2505))
        job.events.append(JobEvent(event_type="completed", payload={"status": "completed"}))
        session.add(job)
        session.commit()
    engine.dispose()
    settings = Settings(data_dir=tmp_path / "data", database_url=f"sqlite:///{database}")
    with TestClient(create_app(settings=settings, excel_runner=FakeExcelRunner())) as client:
        with client.stream("GET", f"/api/excel-jobs/{public_id}/events") as response:
            lines = list(response.iter_lines())
        ids = [int(line[4:]) for line in lines if line.startswith("id: ")]
        assert len(ids) == 2506
        assert any(line == "event: completed" for line in lines[-5:])
        with client.stream(
            "GET", f"/api/excel-jobs/{public_id}/events", headers={"Last-Event-ID": str(ids[-3])}
        ) as response:
            replay = [int(line[4:]) for line in response.iter_lines() if line.startswith("id: ")]
        assert replay == ids[-2:]


def test_running_cancel_is_idempotent_and_preserves_source(tmp_path) -> None:
    runner = FakeExcelRunner(block=True)
    settings = Settings(data_dir=tmp_path / "data", database_url=f"sqlite:///{tmp_path / 'api.db'}")
    with TestClient(create_app(settings=settings, excel_runner=runner)) as client:
        job = upload(client).json()
        for _ in range(100):
            detail = client.get(f"/api/excel-jobs/{job['id']}").json()
            if detail["status"] == "running":
                break
            import time

            time.sleep(0.01)
        first = client.post(f"/api/excel-jobs/{job['id']}/cancel")
        second = client.post(f"/api/excel-jobs/{job['id']}/cancel")
        assert first.status_code == second.status_code == 200
        assert wait_terminal(client, job["id"])["status"] == "cancelled"
        assert sha256(runner.calls[0].source_path) == sha256(FIXTURE)


def test_real_runner_command_invokes_only_employee_entry_and_trusted_arguments(tmp_path) -> None:
    repository = Path(__file__).parents[2]
    request = RunnerRequest(
        public_id="00000000-0000-4000-8000-000000000000",
        source_path=tmp_path / "source.xlsx",
        operation_dir=tmp_path / "operation",
        rules_path=tmp_path / "rules.json",
        knowledge_path=tmp_path / "knowledge.json",
        knowledge_export_id="kx-" + "a" * 32,
        knowledge_payload_sha256="b" * 64,
        knowledge_file_sha256="c" * 64,
    )
    command = build_employee_command(request, repository_root=repository, python_executable="python")
    assert command[:2] == ["python", str(repository / "employee/skills/etsy-performance-listing/scripts/run_task.py")]
    assert command[2:4] == [str(request.source_path), str(request.operation_dir)]
    assert "inspect_workbook.py" not in " ".join(command)
    assert "--knowledge" in command
    assert "--expected-knowledge-export-id" in command
    assert "--expected-knowledge-payload-sha256" in command
    assert "--expected-knowledge-file-sha256" in command


def test_backend_rejects_worker_artifact_that_copies_guarded_competitor_text(tmp_path) -> None:
    raw = "Velvet Vampire Cape for Women Dramatic Gothic Halloween Costume"

    class CopiedRunner(FakeExcelRunner):
        async def run(self, request, emit):
            self.calls.append(request)
            output = request.operation_dir / "generated.xlsx"
            shutil.copyfile(request.source_path, output)
            workbook = load_workbook(output)
            sheet = workbook.active
            for cells in sheet.iter_rows(min_row=1, max_row=20):
                headers = {str(cell.value).strip(): cell.column for cell in cells if cell.value}
                if {"head titles", "SPECIFICATION", "Instructions for buyers"} <= set(headers):
                    header_row = cells[0].row
                    break
            sheet.cell(header_row + 2, headers["head titles"], raw)
            sheet.cell(header_row + 2, headers["SPECIFICATION"], raw)
            sheet.cell(header_row + 2, headers["Instructions for buyers"], raw)
            workbook.save(output)
            return WorkerResult(output, sha256(output))

    settings = Settings(data_dir=tmp_path / "data", database_url=f"sqlite:///{tmp_path / 'api.db'}")
    with TestClient(create_app(settings=settings, excel_runner=CopiedRunner())) as client:
        from app.knowledge.schemas import EvidenceInput
        from datetime import UTC, datetime
        client.app.state.knowledge_service.ingest_evidence(EvidenceInput(
            url="https://www.etsy.com/listing/123/sample", title=raw,
            snapshot="Protected public competitor snapshot.", tags=[],
            source_timestamp=datetime(2026, 8, 10, tzinfo=UTC),
        ))
        job = upload(client).json()
        terminal = wait_terminal(client, job["id"])
        assert terminal["status"] == "failed"
        assert terminal["error"]["code"] == "originality_failed"
        assert raw not in json.dumps(terminal)
        assert client.get(f"/api/excel-jobs/{job['id']}/download").status_code == 409
        workspace = settings.data_dir / "excel-jobs" / job["id"]
        assert not (workspace / "operations").exists() or not list((workspace / "operations").rglob("*.xlsx"))


def test_cleanup_failure_is_persisted_safely_and_job_remains_failed(tmp_path, monkeypatch) -> None:
    settings = Settings(data_dir=tmp_path / "data", database_url=f"sqlite:///{tmp_path / 'api.db'}")
    app = create_app(settings=settings, excel_runner=FakeExcelRunner())

    def fail_cleanup(operation, workspace):
        raise StorageError("cleanup_failed", "private full path and raw evidence")

    monkeypatch.setattr("app.excel_jobs.service.remove_operation_dir", fail_cleanup)
    with TestClient(app) as client:
        job = upload(client).json()
        for _ in range(200):
            terminal = client.get(f"/api/excel-jobs/{job['id']}").json()
            if (terminal.get("error") or {}).get("code") == "cleanup_failed":
                break
            time.sleep(0.01)
        assert terminal["status"] == "failed"
        assert terminal["error"]["code"] == "cleanup_failed"
        assert "private" not in json.dumps(terminal).casefold()

        with app.state.session_factory() as session:
            stored = session.scalar(select(ExcelJob).where(ExcelJob.public_id == job["id"]))
            cleanup_events = [event.payload for event in stored.events if event.event_type == "cleanup_failed"]
        assert cleanup_events == [{"status": "failed", "error": {"code": "cleanup_failed", "message": "Temporary operation cleanup failed."}}]


def test_remove_operation_dir_retries_transient_windows_lock(tmp_path, monkeypatch) -> None:
    workspace = tmp_path / "job"
    operation = workspace / "operations" / "op"
    operation.mkdir(parents=True)
    (operation / "generated.xlsx").write_bytes(b"temporary")
    real_rmtree = shutil.rmtree
    attempts = 0

    def locked_twice(path, *args, **kwargs):
        nonlocal attempts
        attempts += 1
        if attempts < 3:
            raise PermissionError("synthetic lock")
        return real_rmtree(path, *args, **kwargs)

    monkeypatch.setattr(excel_storage.shutil, "rmtree", locked_twice)
    excel_storage.remove_operation_dir(operation, workspace)

    assert attempts == 3
    assert not operation.exists()


def test_artifact_contract_rejects_duplicate_fixed_header(tmp_path) -> None:
    runner = FakeExcelRunner()

    class DuplicateRunner(FakeExcelRunner):
        async def run(self, request, emit):
            result = await super().run(request, emit)
            workbook = load_workbook(result.output_path)
            worksheet = workbook.active
            worksheet.cell(1, worksheet.max_column + 1, "head titles")
            workbook.save(result.output_path)
            return WorkerResult(result.output_path, sha256(result.output_path))

    settings = Settings(data_dir=tmp_path / "data", database_url=f"sqlite:///{tmp_path / 'api.db'}")
    with TestClient(create_app(settings=settings, excel_runner=DuplicateRunner())) as client:
        job = upload(client).json()
        terminal = wait_terminal(client, job["id"])
        assert terminal["status"] == "failed"
        assert terminal["error"]["code"] == "invalid_artifact"


def test_unknown_job_and_invalid_pagination_are_rejected(api) -> None:
    client, _, _, _ = api
    assert client.get("/api/excel-jobs/not-a-uuid").status_code == 422
    assert client.get("/api/excel-jobs/00000000-0000-4000-8000-000000000000").status_code == 404
    assert client.get("/api/excel-jobs", params={"limit": 101}).status_code == 422


def test_worker_artifact_path_escape_is_rejected_without_deleting_external_file(tmp_path) -> None:
    outside = tmp_path / "outside.xlsx"
    shutil.copyfile(FIXTURE, outside)

    class EscapeRunner(FakeExcelRunner):
        async def run(self, request, emit):
            self.calls.append(request)
            await emit({"event": "started"})
            return WorkerResult(outside, sha256(outside))

    settings = Settings(data_dir=tmp_path / "data", database_url=f"sqlite:///{tmp_path / 'api.db'}")
    with TestClient(create_app(settings=settings, excel_runner=EscapeRunner())) as client:
        job = upload(client).json()
        terminal = wait_terminal(client, job["id"])
        assert terminal["status"] == "failed"
        assert terminal["error"]["code"] == "unsafe_path"
        assert outside.exists()
        assert outside.read_bytes() == FIXTURE.read_bytes()


def test_worker_source_mutation_is_detected_and_source_is_never_downloadable(tmp_path) -> None:
    class MutatingRunner(FakeExcelRunner):
        async def run(self, request, emit):
            self.calls.append(request)
            await emit({"event": "started"})
            request.source_path.chmod(stat.S_IWRITE | stat.S_IREAD)
            request.source_path.write_bytes(b"mutated")
            raise RuntimeError("failed after mutation")

    settings = Settings(data_dir=tmp_path / "data", database_url=f"sqlite:///{tmp_path / 'api.db'}")
    with TestClient(create_app(settings=settings, excel_runner=MutatingRunner())) as client:
        job = upload(client).json()
        terminal = wait_terminal(client, job["id"])
        assert terminal["status"] == "failed"
        assert terminal["error"]["code"] == "source_modified"
        assert client.get(f"/api/excel-jobs/{job['id']}/download").status_code == 409


def test_worker_source_deletion_is_detected_without_task_crash(tmp_path) -> None:
    class DeletingRunner(FakeExcelRunner):
        async def run(self, request, emit):
            self.calls.append(request)
            await emit({"event": "started"})
            request.source_path.chmod(stat.S_IWRITE | stat.S_IREAD)
            request.source_path.unlink()
            raise RuntimeError("failed after deleting source")

    settings = Settings(data_dir=tmp_path / "data", database_url=f"sqlite:///{tmp_path / 'api.db'}")
    with TestClient(create_app(settings=settings, excel_runner=DeletingRunner())) as client:
        job = upload(client).json()
        terminal = wait_terminal(client, job["id"])
        assert terminal["status"] == "failed"
        assert terminal["error"]["code"] == "source_modified"


def test_worker_hardlink_to_source_is_not_a_new_artifact(tmp_path) -> None:
    class HardlinkRunner(FakeExcelRunner):
        async def run(self, request, emit):
            self.calls.append(request)
            await emit({"event": "started"})
            output = request.operation_dir / "generated.xlsx"
            request.source_path.chmod(stat.S_IWRITE | stat.S_IREAD)
            try:
                output.hardlink_to(request.source_path)
            except OSError:
                pytest.skip("hard links are unavailable on this filesystem")
            return WorkerResult(output, sha256(output))

    settings = Settings(data_dir=tmp_path / "data", database_url=f"sqlite:///{tmp_path / 'api.db'}")
    with TestClient(create_app(settings=settings, excel_runner=HardlinkRunner())) as client:
        job = upload(client).json()
        terminal = wait_terminal(client, job["id"])
        assert terminal["status"] == "failed"
        assert terminal["error"]["code"] == "invalid_artifact"


def test_worker_cross_job_hardlink_is_rejected(tmp_path) -> None:
    external = tmp_path / "other-job-artifact.xlsx"
    shutil.copyfile(FIXTURE, external)

    class CrossJobHardlinkRunner(FakeExcelRunner):
        async def run(self, request, emit):
            self.calls.append(request)
            await emit({"event": "started"})
            output = request.operation_dir / "generated.xlsx"
            try:
                output.hardlink_to(external)
            except OSError:
                pytest.skip("hard links are unavailable on this filesystem")
            return WorkerResult(output, sha256(output))

    settings = Settings(data_dir=tmp_path / "data", database_url=f"sqlite:///{tmp_path / 'api.db'}")
    with TestClient(create_app(settings=settings, excel_runner=CrossJobHardlinkRunner())) as client:
        job = upload(client).json()
        terminal = wait_terminal(client, job["id"])
        assert terminal["status"] == "failed"
        assert terminal["error"]["code"] == "invalid_artifact"
        assert external.exists()


def test_worker_leaf_symlink_artifact_is_rejected(tmp_path) -> None:
    external = tmp_path / "outside.xlsx"
    shutil.copyfile(FIXTURE, external)
    probe = tmp_path / "symlink-probe.xlsx"
    try:
        probe.symlink_to(external)
        probe.unlink()
    except OSError:
        pytest.skip("symbolic links are unavailable on this filesystem")

    class SymlinkRunner(FakeExcelRunner):
        async def run(self, request, emit):
            self.calls.append(request)
            await emit({"event": "started"})
            output = request.operation_dir / "generated.xlsx"
            output.symlink_to(external)
            return WorkerResult(output, sha256(output))

    settings = Settings(data_dir=tmp_path / "data", database_url=f"sqlite:///{tmp_path / 'api.db'}")
    with TestClient(create_app(settings=settings, excel_runner=SymlinkRunner())) as client:
        job = upload(client).json()
        terminal = wait_terminal(client, job["id"])
        assert terminal["status"] == "failed"
        assert terminal["error"]["code"] == "unsafe_path"
        assert external.exists()


def test_missing_persisted_artifact_download_is_safe_conflict(api) -> None:
    client, _, _, app = api
    job = upload(client).json()
    terminal = wait_terminal(client, job["id"])
    artifact_path = Path(app.state.excel_job_service.download(job["id"])[0])
    artifact_path.unlink()
    response = client.get(f"/api/excel-jobs/{job['id']}/download")
    assert response.status_code in {409, 410}


def test_startup_marks_persisted_queued_and_running_jobs_failed(tmp_path) -> None:
    database = tmp_path / "restart.db"
    engine = create_engine_for_url(f"sqlite:///{database}")
    init_db(engine)
    factory = create_session_factory(engine)
    with factory() as session:
        session.add_all(
            [
                ExcelJob(public_id="00000000-0000-4000-8000-000000000001", source_filename="queued.xlsx", source_sha256="a" * 64, source_size_bytes=1, status=JobStatus.QUEUED),
                ExcelJob(public_id="00000000-0000-4000-8000-000000000002", source_filename="running.xlsx", source_sha256="b" * 64, source_size_bytes=1, status=JobStatus.RUNNING),
            ]
        )
        session.commit()
    engine.dispose()
    settings = Settings(data_dir=tmp_path / "data", database_url=f"sqlite:///{database}")
    with TestClient(create_app(settings=settings, excel_runner=FakeExcelRunner())) as client:
        for public_id in ("00000000-0000-4000-8000-000000000001", "00000000-0000-4000-8000-000000000002"):
            detail = client.get(f"/api/excel-jobs/{public_id}")
            assert detail.status_code == 200
            assert detail.json()["status"] == "failed"
            assert detail.json()["error"]["code"] == "interrupted"


def test_publish_uses_exclusive_random_temp_and_never_overwrites_legacy_hardlink(tmp_path) -> None:
    public_id = "00000000-0000-4000-8000-000000000123"
    workspace = tmp_path / public_id
    operation = workspace / "operations" / "op"
    operation.mkdir(parents=True)
    generated = operation / "generated.xlsx"
    shutil.copyfile(FIXTURE, generated)
    artifact_dir = workspace / "artifacts"
    artifact_dir.mkdir()
    external = tmp_path / "external-protected.bin"
    external.write_bytes(b"must stay unchanged")
    predictable = artifact_dir / f".{public_id}.tmp.xlsx"
    try:
        predictable.hardlink_to(external)
    except OSError:
        pytest.skip("hard links are unavailable on this filesystem")

    published = publish_artifact(
        generated,
        workspace=workspace,
        public_id=public_id,
        expected_sha256=sha256(generated),
    )

    assert published.read_bytes() == FIXTURE.read_bytes()
    assert external.read_bytes() == b"must stay unchanged"
    assert predictable.samefile(external)


def test_publish_rejects_preexisting_destination_without_touching_external_hardlink(tmp_path) -> None:
    public_id = "00000000-0000-4000-8000-000000000124"
    workspace = tmp_path / public_id
    operation = workspace / "operations" / "op"
    operation.mkdir(parents=True)
    generated = operation / "generated.xlsx"
    shutil.copyfile(FIXTURE, generated)
    artifact_dir = workspace / "artifacts"
    artifact_dir.mkdir()
    external = tmp_path / "external-target.bin"
    external.write_bytes(b"protected target")
    destination = artifact_dir / f"etsy-listings-{public_id}.xlsx"
    try:
        destination.hardlink_to(external)
    except OSError:
        pytest.skip("hard links are unavailable on this filesystem")

    with pytest.raises(StorageError, match="destination"):
        publish_artifact(
            generated,
            workspace=workspace,
            public_id=public_id,
            expected_sha256=sha256(generated),
        )

    assert external.read_bytes() == b"protected target"
    assert destination.samefile(external)


def test_publish_detects_staging_replacement_and_never_deletes_attacker_file(tmp_path, monkeypatch) -> None:
    public_id = "00000000-0000-4000-8000-000000000125"
    workspace = tmp_path / public_id
    operation = workspace / "operations" / "op"
    operation.mkdir(parents=True)
    generated = operation / "generated.xlsx"
    shutil.copyfile(FIXTURE, generated)
    external = tmp_path / "race-protected.bin"
    external.write_bytes(b"race protected")
    real_replace = os.replace
    attacked = False

    def replace_with_race(source, destination):
        nonlocal attacked
        source = Path(source)
        if not attacked and source.name.startswith(".publish-"):
            attacked = True
            source.unlink()
            source.hardlink_to(external)
        return real_replace(source, destination)

    monkeypatch.setattr(excel_storage.os, "replace", replace_with_race)
    with pytest.raises(StorageError, match="(staging file|destination) changed"):
        publish_artifact(
            generated,
            workspace=workspace,
            public_id=public_id,
            expected_sha256=sha256(generated),
        )
    assert attacked
    assert external.read_bytes() == b"race protected"


def test_publish_target_replacement_race_never_overwrites_attacker_file(tmp_path, monkeypatch) -> None:
    public_id = "00000000-0000-4000-8000-000000000126"
    workspace = tmp_path / public_id
    operation = workspace / "operations" / "op"
    operation.mkdir(parents=True)
    generated = operation / "generated.xlsx"
    shutil.copyfile(FIXTURE, generated)
    external = tmp_path / "target-race-protected.bin"
    external.write_bytes(b"target race protected")
    real_replace = os.replace
    attacked = False

    def replace_with_race(source, destination):
        nonlocal attacked
        source = Path(source)
        destination = Path(destination)
        if not attacked and source.name.startswith(".publish-") and destination.name == f"etsy-listings-{public_id}.xlsx":
            attacked = True
            destination.unlink()
            destination.hardlink_to(external)
        return real_replace(source, destination)

    monkeypatch.setattr(excel_storage.os, "replace", replace_with_race)
    published = publish_artifact(
        generated,
        workspace=workspace,
        public_id=public_id,
        expected_sha256=sha256(generated),
    )
    destination = workspace / "artifacts" / f"etsy-listings-{public_id}.xlsx"
    assert attacked
    assert published == destination
    assert not destination.samefile(external)
    assert destination.read_bytes() == FIXTURE.read_bytes()
    assert external.read_bytes() == b"target race protected"


def _runner_request(tmp_path: Path, public_id: str) -> RunnerRequest:
    return RunnerRequest(
        public_id=public_id,
        source_path=tmp_path / "source.xlsx",
        operation_dir=tmp_path / "operation",
        rules_path=tmp_path / "rules.json",
        knowledge_path=tmp_path / "knowledge.json",
        knowledge_export_id="kx-" + "a" * 32,
        knowledge_payload_sha256="b" * 64,
        knowledge_file_sha256="c" * 64,
    )


def test_subprocess_runner_stderr_limit_terminates_blocked_worker_promptly(tmp_path, monkeypatch) -> None:
    script = "import sys,time;sys.stderr.write('x'*200000);sys.stderr.flush();time.sleep(30)"
    monkeypatch.setattr(
        "app.excel_jobs.runner.build_employee_command",
        lambda request, repository_root: [sys.executable, "-c", script],
    )
    runner = SubprocessExcelRunner(
        repository_root=tmp_path,
        max_event_bytes=1024,
        cancel_timeout_seconds=0.5,
        worker_timeout_seconds=10,
    )

    async def exercise() -> None:
        async def emit(_event):
            return None

        started = time.monotonic()
        with pytest.raises(WorkerProtocolError, match="stderr"):
            await runner.run(_runner_request(tmp_path, "stderr-limit"), emit)
        assert time.monotonic() - started < 3
        assert not runner._processes

    asyncio.run(exercise())


def test_subprocess_runner_has_bounded_overall_timeout(tmp_path, monkeypatch) -> None:
    monkeypatch.setattr(
        "app.excel_jobs.runner.build_employee_command",
        lambda request, repository_root: [sys.executable, "-c", "import time;time.sleep(30)"],
    )
    runner = SubprocessExcelRunner(
        repository_root=tmp_path,
        max_event_bytes=1024,
        cancel_timeout_seconds=0.5,
        worker_timeout_seconds=0.2,
    )

    async def exercise() -> None:
        started = time.monotonic()
        with pytest.raises(WorkerProtocolError, match="time limit"):
            await runner.run(_runner_request(tmp_path, "overall-timeout"), lambda _: asyncio.sleep(0))
        assert time.monotonic() - started < 3
        assert not runner._processes

    asyncio.run(exercise())


def test_wakeup_subscriptions_are_independent_and_evicted(api) -> None:
    client, _, _, app = api
    service = app.state.excel_job_service
    first = service.subscribe("job-a")
    second = service.subscribe("job-a")
    other = service.subscribe("job-b")
    service._signal("job-a")
    assert first.is_set() and second.is_set()
    assert not other.is_set()
    service.unsubscribe("job-a", first)
    service.unsubscribe("job-a", second)
    service.unsubscribe("job-b", other)
    assert service._wakeups == {}

    for _ in range(3):
        job = upload(client).json()
        wait_terminal(client, job["id"])
        with client.stream("GET", f"/api/excel-jobs/{job['id']}/events") as response:
            list(response.iter_lines())
    assert service._wakeups == {}


def test_legacy_migrated_job_is_visible_in_list_and_detail_api(tmp_path) -> None:
    database = tmp_path / "legacy-api.db"
    engine = create_engine_for_url(f"sqlite:///{database}")
    with engine.begin() as connection:
        connection.execute(text("CREATE TABLE conversations (id INTEGER PRIMARY KEY, title VARCHAR(255) NOT NULL, created_at DATETIME NOT NULL, updated_at DATETIME NOT NULL)"))
        connection.execute(text("CREATE TABLE excel_jobs (id INTEGER PRIMARY KEY, conversation_id INTEGER, source_filename VARCHAR(255) NOT NULL, status VARCHAR(12) NOT NULL, error TEXT, created_at DATETIME NOT NULL, updated_at DATETIME NOT NULL)"))
        connection.execute(text("CREATE TABLE artifacts (id INTEGER PRIMARY KEY, excel_job_id INTEGER NOT NULL, kind VARCHAR(63) NOT NULL, path TEXT NOT NULL, created_at DATETIME NOT NULL)"))
        connection.execute(text("INSERT INTO excel_jobs VALUES (42, NULL, 'old.xlsx', 'running', NULL, CURRENT_TIMESTAMP, CURRENT_TIMESTAMP)"))
    engine.dispose()
    settings = Settings(data_dir=tmp_path / "data", database_url=f"sqlite:///{database}")
    expected_id = str(uuid5(NAMESPACE_URL, "etsy-performance-employee:legacy-excel-job:42"))
    with TestClient(create_app(settings=settings, excel_runner=FakeExcelRunner())) as client:
        listing = client.get("/api/excel-jobs")
        assert listing.status_code == 200
        assert listing.json()["total"] == 1
        assert listing.json()["items"][0]["id"] == expected_id
        detail = client.get(f"/api/excel-jobs/{expected_id}")
        assert detail.status_code == 200
        assert detail.json()["status"] == "failed"
        assert detail.json()["error"]["code"] == "legacy_migrated"
