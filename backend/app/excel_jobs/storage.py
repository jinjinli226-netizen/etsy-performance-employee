from __future__ import annotations

import asyncio
import hashlib
import json
import os
import re
import secrets
import shutil
import stat
import time
import zipfile
import ctypes
from dataclasses import dataclass
from typing import BinaryIO
from pathlib import Path
from uuid import UUID

from fastapi import UploadFile
from openpyxl import load_workbook


OUTPUT_HEADERS = (
    "head titles",
    "13 tags",
    "SPECIFICATION",
    "Category",
    "Instructions for buyers",
)
MAX_UNCOMPRESSED_BYTES = 250 * 1024 * 1024
MAX_ZIP_ENTRIES = 10_000
MAX_ROWS = 20_000
MAX_COLUMNS = 500
UPLOAD_CHUNK_BYTES = 1024 * 1024
_SHA256 = re.compile(r"^[0-9a-f]{64}$")
FILE_ATTRIBUTE_REPARSE_POINT = 0x400
INVALID_FILE_ATTRIBUTES = 0xFFFFFFFF


class StorageError(RuntimeError):
    def __init__(self, code: str, message: str, status_code: int = 422) -> None:
        super().__init__(message)
        self.code = code
        self.status_code = status_code


@dataclass(frozen=True)
class StoredSource:
    workspace: Path
    source_path: Path
    sha256: str
    size_bytes: int


@dataclass(frozen=True)
class KnowledgeTrust:
    path: Path
    export_id: str
    payload_sha256: str
    file_sha256: str


def file_sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for chunk in iter(lambda: stream.read(UPLOAD_CHUNK_BYTES), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _within(path: Path, parent: Path) -> bool:
    try:
        path.relative_to(parent)
        return True
    except ValueError:
        return False


def safe_path(path: Path, parent: Path, *, must_exist: bool = False) -> Path:
    parent_resolved = parent.resolve(strict=True)
    lexical = Path(os.path.abspath(path))
    if not _within(lexical, parent_resolved) or lexical == parent_resolved:
        raise StorageError("unsafe_path", "The operation path is outside its job workspace.")
    current = lexical if must_exist else lexical.parent
    while current != parent_resolved:
        if current.exists() and (current.is_symlink() or _is_reparse_point(current)):
            raise StorageError("unsafe_path", "Links and reparse points are not allowed in job workspaces.")
        current = current.parent
    resolved = path.resolve(strict=must_exist)
    if not _within(resolved, parent_resolved) or resolved == parent_resolved:
        raise StorageError("unsafe_path", "The operation path is outside its job workspace.")
    return resolved


def _is_reparse_point(path: Path) -> bool:
    if os.name != "nt":
        return False
    attributes = ctypes.windll.kernel32.GetFileAttributesW(str(path))
    return attributes not in (-1, INVALID_FILE_ATTRIBUTES) and bool(attributes & FILE_ATTRIBUTE_REPARSE_POINT)


def _validate_regular_unique_file(path: Path) -> None:
    if path.is_symlink() or _is_reparse_point(path) or not path.is_file():
        raise StorageError("invalid_artifact", "The artifact must be a regular file.")
    metadata = path.stat(follow_symlinks=False)
    if metadata.st_nlink != 1:
        raise StorageError("invalid_artifact", "The artifact must not share storage with another file.")


def _validate_xlsx_package(path: Path | BinaryIO) -> None:
    try:
        if hasattr(path, "seek"):
            path.seek(0)
        with zipfile.ZipFile(path) as archive:
            infos = archive.infolist()
            if len(infos) > MAX_ZIP_ENTRIES:
                raise StorageError("unsafe_workbook", "The workbook package has too many entries.")
            names = {info.filename.replace("\\", "/").casefold().lstrip("/") for info in infos}
            if "[content_types].xml" not in names or "xl/workbook.xml" not in names:
                raise StorageError("invalid_workbook", "The file is not a valid XLSX workbook.")
            if any(name.endswith("vbaproject.bin") for name in names):
                raise StorageError("unsupported_workbook", "Macro-enabled workbooks are not supported.", 415)
            total = 0
            for info in infos:
                total += info.file_size
                if total > MAX_UNCOMPRESSED_BYTES:
                    raise StorageError("unsafe_workbook", "The workbook expands beyond the safe limit.")
                if info.file_size > 10 * 1024 * 1024 and info.compress_size and info.file_size / info.compress_size > 200:
                    raise StorageError("unsafe_workbook", "The workbook contains a suspicious archive entry.")
        if hasattr(path, "seek"):
            path.seek(0)
        workbook = load_workbook(path, read_only=True, data_only=False, keep_links=False)
        try:
            for worksheet in workbook.worksheets:
                if worksheet.max_row is None or worksheet.max_column is None:
                    worksheet.calculate_dimension(force=True)
                row_count = worksheet.max_row or 0
                column_count = worksheet.max_column or 0
                if row_count > MAX_ROWS or column_count > MAX_COLUMNS:
                    raise StorageError("unsafe_workbook", "A worksheet exceeds safe dimension limits.")
        finally:
            workbook.close()
    except StorageError:
        raise
    except (OSError, zipfile.BadZipFile, KeyError, ValueError) as exc:
        raise StorageError("invalid_workbook", "The file is not a valid XLSX workbook.") from exc
    except Exception as exc:
        raise StorageError("invalid_workbook", "The file is not a valid XLSX workbook.") from exc


async def store_upload(
    upload: UploadFile,
    *,
    root: Path,
    public_id: UUID,
    max_bytes: int,
) -> StoredSource:
    filename = Path(upload.filename or "").name
    if Path(filename).suffix.casefold() != ".xlsx":
        raise StorageError("unsupported_file_type", "Only .xlsx workbooks are supported.", 415)
    workspace = root / str(public_id)
    if workspace.exists():
        raise StorageError("workspace_exists", "The job workspace already exists.", 409)
    workspace.mkdir(parents=True, exist_ok=False)
    source_dir = workspace / "source"
    source_dir.mkdir()
    source_path = source_dir / "source.xlsx"
    digest = hashlib.sha256()
    size = 0
    try:
        with source_path.open("xb") as stream:
            while True:
                chunk = await upload.read(UPLOAD_CHUNK_BYTES)
                if not chunk:
                    break
                size += len(chunk)
                if size > max_bytes:
                    raise StorageError("upload_too_large", "The workbook exceeds the upload limit.", 413)
                stream.write(chunk)
                digest.update(chunk)
            stream.flush()
            os.fsync(stream.fileno())
        if size == 0:
            raise StorageError("empty_upload", "The uploaded workbook is empty.")
        await asyncio.to_thread(_validate_xlsx_package, source_path)
        os.chmod(source_path, stat.S_IREAD)
        return StoredSource(workspace, source_path, digest.hexdigest(), size)
    except Exception:
        if workspace.exists() and workspace.resolve().parent == root.resolve():
            shutil.rmtree(workspace, ignore_errors=True)
        raise
    finally:
        await upload.close()


def create_operation_dir(workspace: Path, operation_id: str) -> Path:
    operation_root = workspace / "operations"
    operation_root.mkdir(exist_ok=True)
    operation = operation_root / operation_id
    safe_path(operation, workspace)
    operation.mkdir(exist_ok=False)
    return safe_path(operation, workspace, must_exist=True)


def discard_unclaimed_upload(stored: StoredSource, root: Path) -> None:
    root_resolved = root.resolve(strict=True)
    workspace = Path(os.path.abspath(stored.workspace))
    try:
        UUID(workspace.name)
    except ValueError as exc:
        raise StorageError("unsafe_path", "The unclaimed upload workspace is invalid.") from exc
    if workspace.parent != root_resolved or not workspace.exists():
        raise StorageError("unsafe_path", "The unclaimed upload workspace is outside storage.")
    if workspace.is_symlink() or _is_reparse_point(workspace):
        raise StorageError("unsafe_path", "Links are not allowed in upload workspaces.")
    verified_source = safe_path(stored.source_path, workspace, must_exist=True)
    if verified_source != workspace / "source" / "source.xlsx":
        raise StorageError("unsafe_path", "The unclaimed upload source is invalid.")
    def make_writable(function, path, _error) -> None:
        os.chmod(path, stat.S_IWRITE | stat.S_IREAD)
        function(path)

    shutil.rmtree(workspace, onerror=make_writable)


def remove_operation_dir(operation: Path, workspace: Path) -> None:
    try:
        verified = safe_path(operation, workspace, must_exist=True)
    except (StorageError, FileNotFoundError):
        return
    if verified.parent.name != "operations":
        raise StorageError("unsafe_path", "Only operation-scoped temporary directories may be removed.")
    def make_writable_and_retry(function, path, _error) -> None:
        os.chmod(path, stat.S_IWRITE | stat.S_IREAD)
        function(path)

    for attempt in range(8):
        try:
            shutil.rmtree(verified, onerror=make_writable_and_retry)
            return
        except (OSError, PermissionError) as exc:
            if attempt == 7:
                raise StorageError("cleanup_failed", "Temporary operation cleanup failed.") from exc
            time.sleep(0.3 * (attempt + 1))


def ensure_empty_knowledge_export(data_dir: Path) -> KnowledgeTrust:
    export_id = "kx-" + hashlib.sha256(b"empty-active-knowledge-v1").hexdigest()[:32]
    payload = {"schema_version": 1, "export_id": export_id, "issuer": "local-knowledge-pipeline-v1", "records": []}
    payload_sha = hashlib.sha256(
        json.dumps(payload, ensure_ascii=False, sort_keys=True, separators=(",", ":")).encode("utf-8")
    ).hexdigest()
    envelope = {**payload, "content_sha256": payload_sha}
    path = data_dir / "trust" / "empty-active-knowledge.json"
    encoded = json.dumps(envelope, ensure_ascii=False, sort_keys=True, separators=(",", ":")).encode("utf-8")
    if path.exists() and path.read_bytes() != encoded:
        raise StorageError("invalid_trust_store", "The trusted empty knowledge export was modified.", 503)
    if not path.exists():
        temporary = path.with_suffix(".tmp")
        temporary.write_bytes(encoded)
        os.replace(temporary, path)
        os.chmod(path, stat.S_IREAD)
    return KnowledgeTrust(path.resolve(), export_id, payload_sha, hashlib.sha256(encoded).hexdigest())


def ensure_default_rules(workspace: Path) -> Path:
    path = workspace / "rules.json"
    payload = {
        "rule_version": "mvp-default-v2",
        "title_min_words": 3,
        "title_max_words": 14,
        "tag_count": 13,
        "tag_max_chars": 20,
        "description_emoji_sections": 5,
    }
    path.write_text(json.dumps(payload, separators=(",", ":")), encoding="utf-8")
    return safe_path(path, workspace, must_exist=True)


def validate_artifact(path: Path, *, operation_dir: Path, source_path: Path, source_sha256: str) -> tuple[str, int]:
    artifact = safe_path(path, operation_dir, must_exist=True)
    _validate_regular_unique_file(artifact)
    try:
        same_source_file = artifact.samefile(source_path)
    except OSError:
        same_source_file = False
    if same_source_file or artifact.suffix.casefold() != ".xlsx":
        raise StorageError("invalid_artifact", "The worker did not create a new XLSX artifact.")
    if file_sha256(source_path) != source_sha256:
        raise StorageError("source_modified", "The source workbook changed during processing.")
    _validate_xlsx_package(artifact)
    workbook = load_workbook(artifact, read_only=True, data_only=False, keep_links=False)
    counts = {header: 0 for header in OUTPUT_HEADERS}
    try:
        for worksheet in workbook.worksheets:
            # Some valid Excel/WPS exports retain a stale ``A1`` dimension even
            # though the worksheet XML contains the full table. Read-only mode
            # otherwise trusts that hint and silently skips the output headers.
            worksheet.reset_dimensions()
            for row in worksheet.iter_rows(values_only=True):
                for value in row:
                    if isinstance(value, str):
                        normalized = re.sub(r"\s+", " ", value).strip()
                        if normalized in counts:
                            counts[normalized] += 1
    finally:
        workbook.close()
    if any(count != 1 for count in counts.values()):
        raise StorageError("invalid_artifact", "The artifact does not contain exactly one of each fixed output header.")
    return file_sha256(artifact), artifact.stat().st_size


def _identity(metadata: os.stat_result) -> tuple[int, int]:
    return metadata.st_dev, metadata.st_ino


def _validate_open_file(path: Path, metadata: os.stat_result, *, code: str = "invalid_artifact") -> None:
    try:
        lexical = path.lstat()
    except OSError as exc:
        raise StorageError(code, "The artifact path changed during publishing.") from exc
    if (
        path.is_symlink()
        or _is_reparse_point(path)
        or not stat.S_ISREG(lexical.st_mode)
        or not stat.S_ISREG(metadata.st_mode)
        or lexical.st_nlink != 1
        or metadata.st_nlink != 1
        or _identity(lexical) != _identity(metadata)
    ):
        raise StorageError(code, "The artifact path changed during publishing.")


def _open_regular_readonly(path: Path) -> tuple[int, os.stat_result]:
    before = path.lstat()
    if path.is_symlink() or _is_reparse_point(path) or not stat.S_ISREG(before.st_mode) or before.st_nlink != 1:
        raise StorageError("invalid_artifact", "The artifact must be a private regular file.")
    flags = os.O_RDONLY | getattr(os, "O_BINARY", 0) | getattr(os, "O_NOFOLLOW", 0)
    try:
        descriptor = os.open(path, flags)
    except OSError as exc:
        raise StorageError("invalid_artifact", "The artifact could not be opened safely.") from exc
    try:
        opened = os.fstat(descriptor)
        _validate_open_file(path, opened)
        if _identity(before) != _identity(opened):
            raise StorageError("invalid_artifact", "The artifact path changed during opening.")
        return descriptor, opened
    except Exception:
        os.close(descriptor)
        raise


def _exclusive_file(path: Path) -> tuple[int, os.stat_result]:
    flags = os.O_CREAT | os.O_EXCL | os.O_WRONLY | getattr(os, "O_BINARY", 0) | getattr(os, "O_NOFOLLOW", 0)
    descriptor = os.open(path, flags, 0o600)
    opened = os.fstat(descriptor)
    _validate_open_file(path, opened)
    return descriptor, opened


def _unlink_if_identity(path: Path, expected: tuple[int, int]) -> None:
    try:
        current = path.lstat()
        if _identity(current) == expected and not path.is_symlink() and not _is_reparse_point(path):
            path.unlink()
    except OSError:
        pass


def _fsync_directory(path: Path) -> None:
    descriptor: int | None = None
    try:
        descriptor = os.open(path, os.O_RDONLY | getattr(os, "O_BINARY", 0))
        os.fsync(descriptor)
    except OSError:
        pass
    finally:
        if descriptor is not None:
            os.close(descriptor)


def publish_artifact(path: Path, *, workspace: Path, public_id: str, expected_sha256: str) -> Path:
    operation = path.parent.resolve()
    source = safe_path(path, operation, must_exist=True)
    _validate_regular_unique_file(source)
    artifact_dir = workspace / "artifacts"
    artifact_dir.mkdir(exist_ok=True)
    artifact_dir = safe_path(artifact_dir, workspace, must_exist=True)
    destination = artifact_dir / f"etsy-listings-{public_id}.xlsx"
    safe_path(destination, workspace)
    if destination.exists() or destination.is_symlink() or _is_reparse_point(destination):
        raise StorageError("invalid_artifact", "The artifact destination already exists.")

    temporary: Path | None = None
    temporary_identity: tuple[int, int] | None = None
    source_descriptor: int | None = None
    temporary_descriptor: int | None = None
    destination_reservation_descriptor: int | None = None
    destination_reservation_identity: tuple[int, int] | None = None
    destination_identity: tuple[int, int] | None = None
    published_ok = False
    try:
        for _ in range(32):
            candidate = artifact_dir / f".publish-{secrets.token_hex(16)}.tmp"
            safe_path(candidate, workspace)
            try:
                temporary_descriptor, created = _exclusive_file(candidate)
                temporary = candidate
                temporary_identity = _identity(created)
                break
            except FileExistsError:
                continue
        if temporary is None or temporary_descriptor is None or temporary_identity is None:
            raise StorageError("invalid_artifact", "A private artifact staging file could not be created.")

        source_descriptor, source_metadata = _open_regular_readonly(source)
        digest = hashlib.sha256()
        while True:
            chunk = os.read(source_descriptor, UPLOAD_CHUNK_BYTES)
            if not chunk:
                break
            digest.update(chunk)
            view = memoryview(chunk)
            while view:
                written = os.write(temporary_descriptor, view)
                view = view[written:]
        if _identity(os.fstat(source_descriptor)) != _identity(source_metadata):
            raise StorageError("invalid_artifact", "The artifact source changed during copying.")
        _validate_open_file(source, os.fstat(source_descriptor))
        os.fsync(temporary_descriptor)
        os.close(temporary_descriptor)
        temporary_descriptor = None

        verify_descriptor, verified_metadata = _open_regular_readonly(temporary)
        try:
            if _identity(verified_metadata) != temporary_identity:
                raise StorageError("invalid_artifact", "The artifact staging file changed during publishing.")
            with os.fdopen(os.dup(verify_descriptor), "rb") as verify_stream:
                staged_digest = hashlib.sha256()
                for chunk in iter(lambda: verify_stream.read(UPLOAD_CHUNK_BYTES), b""):
                    staged_digest.update(chunk)
                if staged_digest.hexdigest() != expected_sha256 or digest.hexdigest() != expected_sha256:
                    raise StorageError("invalid_artifact", "The published artifact digest changed during copying.")
                _validate_xlsx_package(verify_stream)
            _validate_open_file(temporary, os.fstat(verify_descriptor))
        finally:
            os.close(verify_descriptor)

        try:
            destination_reservation_descriptor, destination_reservation = _exclusive_file(destination)
        except FileExistsError as exc:
            raise StorageError("invalid_artifact", "The artifact destination already exists.") from exc
        destination_reservation_identity = _identity(destination_reservation)
        os.fsync(destination_reservation_descriptor)
        os.close(destination_reservation_descriptor)
        destination_reservation_descriptor = None
        if _identity(destination.lstat()) != destination_reservation_identity:
            raise StorageError("invalid_artifact", "The artifact destination changed during publishing.")
        if _identity(temporary.lstat()) != temporary_identity:
            raise StorageError("invalid_artifact", "The artifact staging file changed during publishing.")

        # Replace only the exclusive reservation that this operation created.
        os.replace(temporary, destination)
        destination_identity = temporary_identity
        if _identity(destination.lstat()) != temporary_identity:
            raise StorageError("invalid_artifact", "The artifact destination changed during publishing.")
        published = safe_path(destination, workspace, must_exist=True)
        published_descriptor, published_metadata = _open_regular_readonly(published)
        try:
            if _identity(published_metadata) != temporary_identity:
                raise StorageError("invalid_artifact", "The artifact staging file changed during publishing.")
            with os.fdopen(os.dup(published_descriptor), "rb") as published_stream:
                published_digest = hashlib.sha256()
                for chunk in iter(lambda: published_stream.read(UPLOAD_CHUNK_BYTES), b""):
                    published_digest.update(chunk)
                if published_digest.hexdigest() != expected_sha256:
                    raise StorageError("invalid_artifact", "The published artifact failed its integrity check.")
                _validate_xlsx_package(published_stream)
            _validate_open_file(published, os.fstat(published_descriptor))
        finally:
            os.close(published_descriptor)
        _fsync_directory(artifact_dir)
        published_ok = True
        return published
    finally:
        for descriptor in (source_descriptor, temporary_descriptor, destination_reservation_descriptor):
            if descriptor is not None:
                try:
                    os.close(descriptor)
                except OSError:
                    pass
        if temporary is not None and temporary_identity is not None:
            _unlink_if_identity(temporary, temporary_identity)
        if not published_ok and destination_identity is not None:
            _unlink_if_identity(destination, destination_identity)
        if not published_ok and destination_reservation_identity is not None:
            _unlink_if_identity(destination, destination_reservation_identity)
