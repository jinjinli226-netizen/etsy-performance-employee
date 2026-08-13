from __future__ import annotations

import hashlib
import json
import os
import stat
import threading
from contextlib import contextmanager
import unicodedata
from dataclasses import dataclass
from pathlib import Path
from typing import Callable
from urllib.parse import urlsplit
from uuid import uuid4

from sqlalchemy import func, select, text
from sqlalchemy.orm import Session, sessionmaker

from app.db.models import AuditEvent, CompetitorEvidence, FeedbackEvent, ImportedEvidenceFingerprint, KnowledgeCandidate, KnowledgePattern, RuleVersion
from app.knowledge.originality import OriginalityGuard
from app.knowledge.promotion import PolicyValidationError, PolicyValidator, PolicyValidatorProtocol, decide_promotion
from app.knowledge.schemas import (
    ActivePatternRead,
    CandidateInput,
    CandidatePage,
    CandidateRead,
    EvidenceInput,
    EvidenceReference,
    KnowledgeKind,
    KnowledgeStatus,
    PatternPage,
    PatternTransitionRead,
)


class KnowledgeNotFoundError(LookupError):
    pass


class KnowledgeConflictError(RuntimeError):
    pass


class KnowledgeValidationError(ValueError):
    pass


class KnowledgeCapacityError(KnowledgeValidationError):
    pass


MAX_GUARD_BYTES = 8 * 1024 * 1024
MAX_GUARD_RECORDS = 500
MAX_GUARD_SHINGLES_PER_RECORD = 30_000


@dataclass(frozen=True)
class KnowledgeTrust:
    path: Path
    export_id: str
    payload_sha256: str
    file_sha256: str


def _canonical_hash(value: object) -> str:
    encoded = json.dumps(value, ensure_ascii=False, sort_keys=True, separators=(",", ":")).encode("utf-8")
    return hashlib.sha256(encoded).hexdigest()


def _snapshot_hash(title: str, snapshot: str, tags: list[str]) -> str:
    def clean(value: str) -> str:
        return " ".join(unicodedata.normalize("NFKC", value).casefold().split())

    return _canonical_hash({"title": clean(title), "snapshot": clean(snapshot), "tags": [clean(tag) for tag in tags]})


def _audit(session: Session, *, actor: str, action: str, entity_type: str, entity_id: str, previous: str | None, new: str | None, trace_id: str | None = None, score: float | None = None) -> None:
    details: dict[str, object] = {"previous": previous, "new": new}
    if trace_id:
        details["trace_id"] = trace_id[:127]
    if score is not None:
        details["max_score"] = round(score, 6)
    session.add(AuditEvent(actor=actor[:127], action=action, entity_type=entity_type, entity_id=entity_id, details=details))


class KnowledgeService:
    def __init__(
        self,
        session_factory: sessionmaker[Session],
        *,
        export_dir: Path,
        originality_threshold: float = 0.72,
        regression_check: Callable[[str, str], bool] | None = None,
        policy_validator: PolicyValidatorProtocol | None = None,
        max_guard_bytes: int = MAX_GUARD_BYTES,
        max_guard_records: int = MAX_GUARD_RECORDS,
    ) -> None:
        self.session_factory = session_factory
        self.export_dir = export_dir
        self.originality = OriginalityGuard(threshold=originality_threshold)
        self.regression_check = regression_check
        self.policy_validator = policy_validator or PolicyValidator()
        self.max_guard_bytes = max(512, min(max_guard_bytes, MAX_GUARD_BYTES))
        self.max_guard_records = max(1, min(max_guard_records, self.originality.max_evidence))
        self._lock = threading.RLock()

    @contextmanager
    def _write_session(self):
        session = self.session_factory()
        try:
            if session.bind is not None and session.bind.dialect.name == "sqlite":
                session.connection().exec_driver_sql("BEGIN IMMEDIATE")
            yield session
            session.commit()
        except Exception:
            session.rollback()
            raise
        finally:
            session.close()

    def ingest_evidence(self, payload: EvidenceInput) -> CompetitorEvidence:
        with self._lock, self._write_session() as session:
            return self._ingest_evidence_in_session(session, payload)

    def capacity_status(self) -> dict[str, int | str]:
        with self.session_factory() as session:
            evidence = list(session.scalars(select(CompetitorEvidence).order_by(CompetitorEvidence.public_id)))
            oversized = sum(
                len(item.snapshot.encode("utf-8")) > 20_000
                or len(item.title.encode("utf-8")) > 500
                for item in evidence
            )
            status = "ready"
            if len(evidence) > self.max_guard_records or oversized:
                status = "exceeded"
            else:
                try:
                    self._guard_envelope(evidence)
                except KnowledgeCapacityError:
                    status = "exceeded"
            payload = {
                "status": status,
                "evidence_count": len(evidence),
                "record_limit": self.max_guard_records,
                "oversized_count": oversized,
            }
            session.execute(text(
                "INSERT INTO knowledge_capacity_state "
                "(id,status,evidence_count,record_limit,oversized_count,updated_at) "
                "VALUES (1,:status,:evidence_count,:record_limit,:oversized_count,CURRENT_TIMESTAMP) "
                "ON CONFLICT(id) DO UPDATE SET status=excluded.status, "
                "evidence_count=excluded.evidence_count, record_limit=excluded.record_limit, "
                "oversized_count=excluded.oversized_count, updated_at=CURRENT_TIMESTAMP"
            ), payload)
            session.commit()
            return payload

    def require_capacity_ready(self) -> None:
        if self.capacity_status()["status"] != "ready":
            raise KnowledgeCapacityError("knowledge capacity exceeded")

    def _ingest_evidence_in_session(self, session: Session, payload: EvidenceInput) -> CompetitorEvidence:
        if len(payload.snapshot.encode("utf-8")) > 20_000 or len(payload.title.encode("utf-8")) > 500:
            raise KnowledgeCapacityError("evidence guard capacity exceeded")
        parsed = urlsplit(payload.url)
        listing_id = parsed.path.split("/")[2]
        canonical = f"https://www.etsy.com/listing/{listing_id}"
        snapshot_payload = {"title": payload.title, "snapshot": payload.snapshot, "tags": payload.tags}
        raw_payload = {"url": canonical, **snapshot_payload}
        digest = _canonical_hash(raw_payload)
        snapshot_digest = _snapshot_hash(payload.title, payload.snapshot, payload.tags)
        existing = session.scalar(
            select(CompetitorEvidence).where(
                CompetitorEvidence.source_key == f"etsy-listing:{listing_id}",
                CompetitorEvidence.content_hash == digest,
            )
        )
        if existing:
            return existing
        record = CompetitorEvidence(
            public_id="ev-" + uuid4().hex,
            canonical_url=canonical,
            source_key=f"etsy-listing:{listing_id}",
            title=payload.title,
            snapshot=payload.snapshot,
            tags=payload.tags,
            source_timestamp=payload.source_timestamp,
            content_hash=digest,
            snapshot_hash=snapshot_digest,
        )
        current = list(session.scalars(select(CompetitorEvidence).order_by(CompetitorEvidence.public_id)))
        self._guard_envelope([*current, record])
        session.add(record)
        session.flush()
        _audit(session, actor="employee", action="evidence_ingested", entity_type="evidence", entity_id=record.public_id, previous=None, new="evidence_only")
        return record

    def _validated_evidence(self, session: Session, payload: CandidateInput) -> list[CompetitorEvidence]:
        ids = list(dict.fromkeys(reference.evidence_id for reference in payload.evidence_refs))
        if not ids:
            return []
        records = list(session.scalars(select(CompetitorEvidence).where(CompetitorEvidence.public_id.in_(ids))))
        by_id = {record.public_id: record for record in records}
        if set(by_id) != set(ids):
            raise KnowledgeValidationError("candidate references unknown evidence")
        for reference in payload.evidence_refs:
            record = by_id[reference.evidence_id]
            if record.source_timestamp != reference.source_timestamp:
                raise KnowledgeValidationError("candidate evidence timestamp mismatch")
        return records

    def _raw_similarity(self, abstract: str, records: list[CompetitorEvidence]):
        evidence = [(record.public_id, f"{record.title} {record.snapshot} {' '.join(record.tags)}") for record in records]
        return self.originality.check_texts([abstract], evidence)

    @staticmethod
    def _independent_count(records: list[CompetitorEvidence]) -> int:
        return min(len({record.source_key for record in records}), len({record.snapshot_hash for record in records}))

    def ingest_candidate(
        self,
        payload: CandidateInput,
        *,
        actor: str,
        trace_id: str,
        conversation_id: int | None = None,
        message_id: int | None = None,
    ) -> KnowledgeCandidate:
        try:
            with self._lock, self._write_session() as session:
                return self._ingest_candidate_in_session(
                    session, payload, actor=actor, trace_id=trace_id,
                    conversation_id=conversation_id, message_id=message_id,
                )
        except KnowledgeValidationError as exc:
            if str(exc) == "candidate is overly similar to raw evidence":
                with self._lock, self._write_session() as session:
                    _audit(session, actor=actor, action="candidate_rejected_similarity", entity_type="candidate", entity_id="pending", previous=None, new="rejected", trace_id=trace_id)
            raise

    def _ingest_candidate_in_session(
        self, session: Session, payload: CandidateInput, *, actor: str, trace_id: str,
        conversation_id: int | None, message_id: int | None,
    ) -> KnowledgeCandidate:
        evidence = self._validated_evidence(session, payload)
        similarity = self._raw_similarity(payload.abstract, evidence)
        if not similarity.passed:
            raise KnowledgeValidationError("candidate is overly similar to raw evidence")
        base_rule, base_revision = self._active_token(session, payload.kind)
        candidate = KnowledgeCandidate(
            public_id="kc-" + uuid4().hex,
            title=payload.abstract[:255],
            proposal={"kind": payload.kind, "abstract": payload.abstract, "confidence": payload.confidence},
            kind=payload.kind,
            abstract_summary=payload.abstract,
            confidence=payload.confidence,
            evidence_ids=[record.public_id for record in evidence],
            source_timestamps={record.public_id: record.source_timestamp.isoformat() for record in evidence},
            conversation_id=conversation_id,
            message_id=message_id,
            trace_id=trace_id[:127],
            base_active_rule_public_id=base_rule,
            base_pattern_revision=base_revision,
            status=KnowledgeStatus.PROPOSED,
        )
        session.add(candidate)
        session.flush()
        _audit(session, actor=actor, action="candidate_proposed", entity_type="candidate", entity_id=candidate.public_id, previous=None, new=KnowledgeStatus.PROPOSED.value, trace_id=trace_id)
        self._maybe_auto_activate(session, candidate, evidence, actor=actor)
        return candidate

    def ingest_learning_batch(
        self,
        payload: dict,
        *,
        allowed_urls: frozenset[str],
        actor: str,
        trace_id: str,
        conversation_id: int | None,
        message_id: int | None,
    ) -> list[KnowledgeCandidate]:
        evidence_inputs = [EvidenceInput.model_validate(item) for item in payload.get("evidence_items", [])]
        candidate_items = payload.get("candidates", [])
        if len(evidence_inputs) > 5 or len(candidate_items) > 5:
            raise KnowledgeValidationError("learning batch exceeds its limit")
        prepared_urls: dict[str, EvidenceInput] = {}
        for evidence in evidence_inputs:
            listing_id = urlsplit(evidence.url).path.split("/")[2]
            canonical = f"https://www.etsy.com/listing/{listing_id}"
            if canonical not in allowed_urls:
                raise KnowledgeValidationError("unapproved evidence URL")
            prepared_urls[canonical] = evidence
        allowed_kinds = {member.value for member in KnowledgeKind}
        for item in candidate_items:
            if item.get("kind") not in allowed_kinds:
                raise KnowledgeValidationError("unknown knowledge kind")
            CandidateInput(kind=item["kind"], abstract=item["summary"], confidence=item["confidence"], evidence_refs=[])
            urls, ids = item.get("evidence_urls", []), item.get("evidence_ids", [])
            if not urls and not ids:
                raise KnowledgeValidationError("learning candidates require operation-bound evidence")
            for url in urls:
                match = __import__("re").search(r"/listing/([0-9]+)", url)
                canonical = f"https://www.etsy.com/listing/{match.group(1)}" if match else ""
                if canonical not in prepared_urls or canonical not in allowed_urls:
                    raise KnowledgeValidationError("candidate evidence is not operation-bound")
        created: list[KnowledgeCandidate] = []
        with self._lock, self._write_session() as session:
            by_url: dict[str, CompetitorEvidence] = {}
            for canonical, evidence in prepared_urls.items():
                by_url[canonical] = self._ingest_evidence_in_session(session, evidence)
            for item in candidate_items:
                records = []
                for url in item.get("evidence_urls", []):
                    listing_id = __import__("re").search(r"/listing/([0-9]+)", url).group(1)
                    records.append(by_url[f"https://www.etsy.com/listing/{listing_id}"])
                ids = item.get("evidence_ids", [])
                if ids:
                    existing = list(session.scalars(select(CompetitorEvidence).where(CompetitorEvidence.public_id.in_(set(ids)))))
                    expected = {"etsy-listing:" + url.split("/listing/", 1)[1] for url in allowed_urls}
                    if len(existing) != len(set(ids)) or any(record.source_key not in expected for record in existing):
                        raise KnowledgeValidationError("evidence is not bound to the learning operation")
                    records.extend(existing)
                candidate_payload = CandidateInput(
                    kind=item["kind"], abstract=item["summary"], confidence=item["confidence"],
                    evidence_refs=[EvidenceReference(evidence_id=record.public_id, source_timestamp=record.source_timestamp) for record in records],
                )
                created.append(self._ingest_candidate_in_session(
                    session, candidate_payload, actor=actor, trace_id=trace_id,
                    conversation_id=conversation_id, message_id=message_id,
                ))
        return created

    def audit_candidate_event_rejection(self, *, actor: str, trace_id: str, reason: str) -> None:
        allowed = {"invalid_schema", "unknown_evidence", "timestamp_mismatch", "unsafe_content", "ingestion_failed", "learning_mode_required", "envelope_invalid"}
        safe_reason = reason if reason in allowed else "ingestion_failed"
        with self.session_factory() as session:
            _audit(session, actor=actor, action="candidate_event_rejected", entity_type="candidate", entity_id="untrusted-event", previous=None, new=safe_reason, trace_id=trace_id)
            session.commit()

    def audit_learning_event_rejection(self, *, actor: str, trace_id: str) -> None:
        with self.session_factory() as session:
            _audit(session, actor=actor, action="learning_event_rejected", entity_type="learning", entity_id="untrusted-event", previous=None, new="learning_mode_required", trace_id=trace_id)
            session.commit()

    def _support_counts(self, session: Session, candidate: KnowledgeCandidate, evidence: list[CompetitorEvidence] | None = None) -> tuple[int, int]:
        if evidence is None:
            ids = candidate.evidence_ids or []
            evidence = list(session.scalars(select(CompetitorEvidence).where(CompetitorEvidence.public_id.in_(ids)))) if ids else []
        edits = list(session.execute(select(FeedbackEvent.feedback_id, FeedbackEvent.row_id).where(FeedbackEvent.knowledge_candidate_id == candidate.id, FeedbackEvent.accepted.is_(True))))
        valid_edits = len({(feedback_id, row_id) for feedback_id, row_id in edits if feedback_id and row_id})
        return self._independent_count(evidence), valid_edits

    def _constraints_pass(self, candidate: KnowledgeCandidate, evidence: list[CompetitorEvidence]) -> tuple[bool, str]:
        try:
            CandidateInput(
                kind=candidate.kind,
                abstract=candidate.abstract_summary,
                confidence=candidate.confidence,
                evidence_refs=[],
            )
        except (TypeError, ValueError):
            return False, "candidate_not_sanitized"
        try:
            self.policy_validator.validate(candidate.kind or "", candidate.abstract_summary or "")
        except PolicyValidationError as exc:
            return False, str(exc) or "policy_failed"
        except Exception:
            return False, "policy_unavailable"
        if self.regression_check is not None:
            try:
                if not self.regression_check(candidate.kind or "", candidate.abstract_summary or ""):
                    return False, "regression_failed"
            except Exception:
                return False, "regression_unavailable"
        if not self._raw_similarity(candidate.abstract_summary or "", evidence).passed:
            return False, "raw_similarity"
        return True, "passed"

    def _maybe_auto_activate(self, session: Session, candidate: KnowledgeCandidate, evidence: list[CompetitorEvidence], *, actor: str) -> None:
        independent, edits = self._support_counts(session, candidate, evidence)
        constraints, reason = self._constraints_pass(candidate, evidence)
        decision = decide_promotion(
            kind=candidate.kind or "",
            confidence=candidate.confidence or 0,
            independent_evidence=independent,
            accepted_edits=edits,
            hard_conflict=reason.startswith("policy_"),
            regression_passed=reason not in {"regression_failed", "regression_unavailable"},
            originality_passed=reason != "raw_similarity",
        )
        if constraints and decision.eligible:
            self._activate(session, candidate, actor="system:auto")

    def _activate(self, session: Session, candidate: KnowledgeCandidate, *, actor: str) -> KnowledgePattern:
        if candidate.status is KnowledgeStatus.REJECTED:
            raise KnowledgeConflictError("rejected candidates cannot be approved")
        existing = session.scalar(select(KnowledgePattern).where(KnowledgePattern.kind == candidate.kind))
        if candidate.status is KnowledgeStatus.ACTIVE and existing:
            return existing
        current_rule, current_revision = self._active_token(session, candidate.kind or "")
        if (
            current_rule != candidate.base_active_rule_public_id
            or current_revision != candidate.base_pattern_revision
        ):
            raise KnowledgeConflictError("candidate is stale and must be rebased")
        ids = candidate.evidence_ids or []
        evidence = list(session.scalars(select(CompetitorEvidence).where(CompetitorEvidence.public_id.in_(ids)))) if ids else []
        passed, reason = self._constraints_pass(candidate, evidence)
        if not passed:
            raise KnowledgeConflictError(reason)
        pattern = existing or KnowledgePattern(
            public_id=str(uuid4()),
            source_candidate=candidate,
            name=candidate.kind or candidate.public_id or str(candidate.id),
            kind=candidate.kind,
            pattern={},
            status=KnowledgeStatus.ACTIVE,
        )
        if existing is None:
            session.add(pattern)
            session.flush()
        current = session.scalar(select(func.max(RuleVersion.sequence)).where(RuleVersion.pattern_id == pattern.id)) or 0
        for version in session.scalars(select(RuleVersion).where(RuleVersion.pattern_id == pattern.id, RuleVersion.status == KnowledgeStatus.ACTIVE)):
            version.status = KnowledgeStatus.ROLLED_BACK
        previous_candidate = pattern.source_candidate
        if previous_candidate is not None and previous_candidate.id != candidate.id:
            previous_candidate.status = KnowledgeStatus.ROLLED_BACK
            previous_candidate.revision += 1
            _audit(
                session,
                actor=actor,
                action="candidate_superseded",
                entity_type="candidate",
                entity_id=previous_candidate.public_id or str(previous_candidate.id),
                previous=KnowledgeStatus.ACTIVE.value,
                new=KnowledgeStatus.ROLLED_BACK.value,
                trace_id=previous_candidate.trace_id,
            )
        public_version = str(uuid4())
        rules = {"kind": candidate.kind, "abstract": candidate.abstract_summary}
        version = RuleVersion(
            public_id=public_version,
            pattern=pattern,
            candidate=candidate,
            sequence=current + 1,
            version=f"knowledge-{pattern.id}-v{current + 1}",
            rules=rules,
            status=KnowledgeStatus.ACTIVE,
        )
        session.add(version)
        previous = candidate.status.value
        candidate.status = KnowledgeStatus.ACTIVE
        candidate.revision += 1
        pattern.source_candidate = candidate
        pattern.abstract_summary = candidate.abstract_summary
        pattern.pattern = rules
        pattern.status = KnowledgeStatus.ACTIVE
        pattern.revision += 1
        _audit(session, actor=actor, action="candidate_activated", entity_type="candidate", entity_id=candidate.public_id or str(candidate.id), previous=previous, new=KnowledgeStatus.ACTIVE.value, trace_id=candidate.trace_id)
        return pattern

    @staticmethod
    def _active_token(session: Session, kind: str) -> tuple[str | None, int | None]:
        pattern = session.scalar(select(KnowledgePattern).where(KnowledgePattern.kind == kind))
        if pattern is None:
            return None, None
        version = session.scalar(
            select(RuleVersion).where(
                RuleVersion.pattern_id == pattern.id,
                RuleVersion.status == KnowledgeStatus.ACTIVE,
            )
        )
        return (version.public_id if version is not None else None), pattern.revision

    def record_accepted_edit(self, candidate_id: int, *, feedback_id: str, row_id: str) -> KnowledgeCandidate:
        if not feedback_id or not row_id or len(feedback_id) > 128 or len(row_id) > 128:
            raise KnowledgeValidationError("invalid feedback reference")
        with self._lock, self._write_session() as session:
            candidate = session.get(KnowledgeCandidate, candidate_id)
            if candidate is None:
                raise KnowledgeNotFoundError
            duplicate = session.scalar(select(FeedbackEvent).where(FeedbackEvent.knowledge_candidate_id == candidate_id, ((FeedbackEvent.feedback_id == feedback_id) | (FeedbackEvent.row_id == row_id))))
            if duplicate is None:
                session.add(FeedbackEvent(public_id=str(uuid4()), knowledge_candidate_id=candidate_id, feedback_id=feedback_id, row_id=row_id, accepted=True, event_type="accepted_edit", payload={}))
                session.flush()
            evidence = list(session.scalars(select(CompetitorEvidence).where(CompetitorEvidence.public_id.in_(candidate.evidence_ids or [])))) if candidate.evidence_ids else []
            if candidate.status is KnowledgeStatus.PROPOSED:
                self._maybe_auto_activate(session, candidate, evidence, actor="feedback")
            return candidate

    def get_candidate(self, candidate_id: int) -> KnowledgeCandidate:
        with self.session_factory() as session:
            candidate = session.get(KnowledgeCandidate, candidate_id)
            if candidate is None:
                raise KnowledgeNotFoundError
            return candidate

    def approve_candidate(self, candidate_id: int, *, actor: str) -> KnowledgePattern:
        with self._lock, self._write_session() as session:
            candidate = session.get(KnowledgeCandidate, candidate_id)
            if candidate is None:
                raise KnowledgeNotFoundError
            pattern = self._activate(session, candidate, actor=actor)
            return pattern

    def reject_candidate(self, candidate_id: int, *, actor: str) -> KnowledgeCandidate:
        with self._lock, self._write_session() as session:
            candidate = session.get(KnowledgeCandidate, candidate_id)
            if candidate is None:
                raise KnowledgeNotFoundError
            if candidate.status is KnowledgeStatus.ACTIVE:
                raise KnowledgeConflictError("active candidates cannot be rejected")
            if candidate.status is KnowledgeStatus.REJECTED:
                return candidate
            previous = candidate.status.value
            candidate.status = KnowledgeStatus.REJECTED
            candidate.revision += 1
            _audit(session, actor=actor, action="candidate_rejected", entity_type="candidate", entity_id=candidate.public_id or str(candidate.id), previous=previous, new=KnowledgeStatus.REJECTED.value)
            return candidate

    def rollback_pattern(
        self, pattern_id: int, *, actor: str, target_version_id: str | None = None,
        expected_rule_version: str | None = None,
    ) -> PatternTransitionRead:
        with self._lock, self._write_session() as session:
            pattern = session.get(KnowledgePattern, pattern_id)
            if pattern is None:
                raise KnowledgeNotFoundError
            versions = list(session.scalars(select(RuleVersion).where(RuleVersion.pattern_id == pattern.id).order_by(RuleVersion.sequence)))
            active = next((version for version in reversed(versions) if version.status is KnowledgeStatus.ACTIVE), None)
            if expected_rule_version is not None and (active is None or active.version != expected_rule_version):
                raise KnowledgeConflictError("rollback request is stale")
            eligible = [version for version in versions if active and version.sequence < active.sequence]
            target = next((version for version in eligible if version.public_id == target_version_id), None) if target_version_id else (eligible[-1] if eligible else None)
            if target is None or active is None or target.id == active.id:
                raise KnowledgeConflictError("no safe previous version exists")
            active.status = KnowledgeStatus.ROLLED_BACK
            active_candidate = active.candidate
            if active_candidate is not None:
                active_candidate.status = KnowledgeStatus.ROLLED_BACK
                active_candidate.revision += 1
                _audit(
                    session,
                    actor=actor,
                    action="candidate_superseded",
                    entity_type="candidate",
                    entity_id=active_candidate.public_id or str(active_candidate.id),
                    previous=KnowledgeStatus.ACTIVE.value,
                    new=KnowledgeStatus.ROLLED_BACK.value,
                    trace_id=active_candidate.trace_id,
                )
            sequence = versions[-1].sequence + 1
            rules = {**target.rules, "rollback_of": target.public_id}
            replacement = RuleVersion(public_id=str(uuid4()), pattern=pattern, candidate=target.candidate, sequence=sequence, version=f"knowledge-{pattern.id}-v{sequence}", rules=rules, status=KnowledgeStatus.ACTIVE)
            session.add(replacement)
            if target.candidate is not None:
                target.candidate.status = KnowledgeStatus.ACTIVE
                target.candidate.revision += 1
                pattern.source_candidate = target.candidate
                _audit(
                    session,
                    actor=actor,
                    action="candidate_restored",
                    entity_type="candidate",
                    entity_id=target.candidate.public_id or str(target.candidate.id),
                    previous=KnowledgeStatus.ROLLED_BACK.value,
                    new=KnowledgeStatus.ACTIVE.value,
                    trace_id=target.candidate.trace_id,
                )
            pattern.pattern = {key: value for key, value in target.rules.items() if key != "rollback_of"}
            pattern.abstract_summary = target.rules["abstract"]
            pattern.revision += 1
            _audit(session, actor=actor, action="pattern_rolled_back", entity_type="pattern", entity_id=pattern.public_id or str(pattern.id), previous=active.public_id, new=replacement.public_id)
            return self._transition(pattern, replacement)

    @staticmethod
    def _transition(pattern: KnowledgePattern, version: RuleVersion | None = None) -> PatternTransitionRead:
        if version is None:
            version = next(item for item in reversed(pattern.rule_versions) if item.status is KnowledgeStatus.ACTIVE)
        return PatternTransitionRead(id=pattern.id, public_id=pattern.public_id or str(pattern.id), kind=pattern.kind or pattern.name, abstract=pattern.abstract_summary or "", rule_version=version.version, status=pattern.status)

    def generation_patterns(self) -> list[dict[str, str]]:
        with self.session_factory() as session:
            patterns = list(session.scalars(select(KnowledgePattern).where(KnowledgePattern.status == KnowledgeStatus.ACTIVE).order_by(KnowledgePattern.kind, KnowledgePattern.id)))
            output = []
            for pattern in patterns:
                versions = list(session.scalars(select(RuleVersion).where(RuleVersion.pattern_id == pattern.id, RuleVersion.status == KnowledgeStatus.ACTIVE)))
                if len(versions) != 1 or pattern.source_candidate_id is None or versions[0].knowledge_candidate_id != pattern.source_candidate_id:
                    raise KnowledgeValidationError("invalid active rule lineage")
                version = versions[0]
                if version:
                    rules = version.rules
                    if not isinstance(rules, dict) or set(rules) - {"kind", "abstract", "rollback_of"}:
                        raise KnowledgeValidationError("invalid active rule snapshot")
                    try:
                        validated = CandidateInput(kind=rules.get("kind"), abstract=rules.get("abstract"), confidence=1, evidence_refs=[])
                        self.policy_validator.validate(validated.kind, validated.abstract)
                    except Exception as exc:
                        raise KnowledgeValidationError("invalid active rule snapshot") from exc
                    output.append({"id": "rec-" + hashlib.sha256((pattern.public_id or str(pattern.id)).encode()).hexdigest()[:16], "kind": validated.kind, "abstract": validated.abstract, "rule_version": version.version})
            return output

    def list_active(self, *, limit: int, offset: int, kind: str | None = None) -> PatternPage:
        records = self.generation_patterns()
        if kind:
            records = [item for item in records if item["kind"] == kind]
        return PatternPage(items=[ActivePatternRead(**item) for item in records[offset : offset + limit]], total=len(records), limit=limit, offset=offset)

    def list_candidates(self, *, limit: int, offset: int, status: KnowledgeStatus | None = None) -> CandidatePage:
        with self.session_factory() as session:
            query = select(KnowledgeCandidate)
            count_query = select(func.count()).select_from(KnowledgeCandidate)
            if status:
                query = query.where(KnowledgeCandidate.status == status)
                count_query = count_query.where(KnowledgeCandidate.status == status)
            total = session.scalar(count_query) or 0
            candidates = list(session.scalars(query.order_by(KnowledgeCandidate.created_at.desc(), KnowledgeCandidate.id.desc()).limit(limit).offset(offset)))
            items = []
            for candidate in candidates:
                independent, edits = self._support_counts(session, candidate)
                items.append(CandidateRead(id=candidate.id, public_id=candidate.public_id or str(candidate.id), kind=candidate.kind or "legacy", abstract=candidate.abstract_summary or "Legacy candidate pending reprocessing.", confidence=candidate.confidence or 0, evidence_count=independent, accepted_edit_count=edits, status=candidate.status, created_at=candidate.created_at, updated_at=candidate.updated_at))
            return CandidatePage(items=items, total=total, limit=limit, offset=offset)

    def export_active_knowledge(self, path: Path) -> KnowledgeTrust:
        records = []
        for pattern in self.generation_patterns():
            record_payload = {"id": pattern["id"], "status": "active", "approved": True, "abstract": pattern["abstract"]}
            records.append({**record_payload, "content_sha256": _canonical_hash(record_payload)})
        identity = _canonical_hash(records)
        export_id = "kx-" + identity[:32]
        payload = {"schema_version": 1, "export_id": export_id, "issuer": "local-knowledge-pipeline-v1", "records": records}
        envelope = {**payload, "content_sha256": _canonical_hash(payload)}
        encoded = json.dumps(envelope, ensure_ascii=False, sort_keys=True, separators=(",", ":")).encode("utf-8")
        path.parent.mkdir(parents=True, exist_ok=True)
        temporary = path.with_name(path.name + ".tmp-" + uuid4().hex)
        temporary.write_bytes(encoded)
        if path.exists():
            os.chmod(path, stat.S_IWRITE | stat.S_IREAD)
        os.replace(temporary, path)
        os.chmod(path, stat.S_IREAD)
        return KnowledgeTrust(path.resolve(), export_id, envelope["content_sha256"], hashlib.sha256(encoded).hexdigest())

    def export_evidence_guard(self, path: Path) -> KnowledgeTrust:
        with self.session_factory() as session:
            evidence = list(session.scalars(select(CompetitorEvidence).order_by(CompetitorEvidence.public_id)))
            imported = list(session.scalars(select(ImportedEvidenceFingerprint).order_by(ImportedEvidenceFingerprint.public_id)))
        envelope, encoded = self._guard_envelope(evidence, imported)
        export_id = envelope["export_id"]
        path.parent.mkdir(parents=True, exist_ok=True)
        temporary = path.with_name(path.name + ".tmp-" + uuid4().hex)
        temporary.write_bytes(encoded)
        if path.exists():
            os.chmod(path, stat.S_IWRITE | stat.S_IREAD)
        os.replace(temporary, path)
        os.chmod(path, stat.S_IREAD)
        return KnowledgeTrust(path.resolve(), export_id, envelope["content_sha256"], hashlib.sha256(encoded).hexdigest())

    def _guard_envelope(self, evidence: list[CompetitorEvidence], imported: list[ImportedEvidenceFingerprint] | None = None) -> tuple[dict[str, object], bytes]:
        imported = imported or []
        if len(evidence) + len(imported) > self.max_guard_records:
            raise KnowledgeCapacityError("evidence guard capacity exceeded")
        records = []
        for item in sorted(evidence, key=lambda record: record.public_id):
            shingles = self.originality.fingerprint_texts([item.title, item.snapshot, *item.tags])
            if len(shingles) > MAX_GUARD_SHINGLES_PER_RECORD:
                raise KnowledgeCapacityError("evidence guard capacity exceeded")
            record_payload = {"id": item.public_id, "shingles": shingles}
            records.append({**record_payload, "content_sha256": _canonical_hash(record_payload)})
        existing = {record["id"] for record in records}
        for item in imported:
            if item.public_id in existing:
                continue
            if len(item.shingles) > MAX_GUARD_SHINGLES_PER_RECORD:
                raise KnowledgeCapacityError("evidence guard capacity exceeded")
            record_payload = {"id": item.public_id, "shingles": sorted(item.shingles)}
            records.append({**record_payload, "content_sha256": _canonical_hash(record_payload)})
        records.sort(key=lambda record: record["id"])
        identity = _canonical_hash(records)
        export_id = "eg-" + identity[:32]
        threshold = min([self.originality.threshold, *(item.threshold for item in imported)])
        payload = {"schema_version": 1, "export_id": export_id, "issuer": "local-evidence-guard-v1", "threshold": threshold, "records": records}
        envelope = {**payload, "content_sha256": _canonical_hash(payload)}
        encoded = json.dumps(envelope, ensure_ascii=False, sort_keys=True, separators=(",", ":")).encode("utf-8")
        if len(encoded) > self.max_guard_bytes:
            raise KnowledgeCapacityError("evidence guard capacity exceeded")
        return envelope, encoded

    def resolve_bound_evidence(self, evidence_ids: list[str], allowed_urls: frozenset[str]) -> list[CompetitorEvidence]:
        if not evidence_ids:
            return []
        with self.session_factory() as session:
            records = list(session.scalars(select(CompetitorEvidence).where(CompetitorEvidence.public_id.in_(set(evidence_ids)))))
            expected_keys = {"etsy-listing:" + url.split("/listing/", 1)[1] for url in allowed_urls}
            if len(records) != len(set(evidence_ids)) or any(record.source_key not in expected_keys for record in records):
                raise KnowledgeValidationError("evidence is not bound to the learning operation")
            session.expunge_all()
            return records

    def validate_generated_workbook(self, path: Path) -> None:
        from openpyxl import load_workbook

        fixed = {
            "head titles": "head_titles",
            "SPECIFICATION": "specification",
            "Instructions for buyers": "instructions_for_buyers",
        }
        with self.session_factory() as session:
            evidence = [
                (item.public_id, self.originality.fingerprint_texts([item.title, item.snapshot, *item.tags]))
                for item in session.scalars(select(CompetitorEvidence).order_by(CompetitorEvidence.public_id))
            ]
        if not evidence:
            return
        try:
            workbook = load_workbook(path, read_only=True, data_only=False)
            try:
                for sheet in workbook.worksheets:
                    header_row, columns = None, {}
                    for row in sheet.iter_rows(min_row=1, max_row=min(sheet.max_row, 200)):
                        found = {str(cell.value).strip(): cell.column for cell in row if isinstance(cell.value, str)}
                        if set(fixed).issubset(found):
                            header_row, columns = row[0].row, found
                            break
                    if header_row is None:
                        continue
                    for row_number in range(header_row + 1, sheet.max_row + 1):
                        generated = {field: sheet.cell(row_number, columns[header]).value or "" for header, field in fixed.items()}
                        result = self.originality.check_fingerprints(generated.values(), evidence)
                        if not result.passed:
                            raise KnowledgeValidationError(
                                json.dumps({"code": "originality_failed", "score": result.max_score, "evidence_id": result.evidence_id}, separators=(",", ":"))
                            )
            finally:
                workbook.close()
        except KnowledgeValidationError:
            raise
        except Exception as exc:
            raise KnowledgeValidationError("generated workbook originality validation failed") from exc
