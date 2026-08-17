from __future__ import annotations

import hashlib
import html as html_module
import json
import re
import unicodedata
from dataclasses import dataclass
from datetime import datetime
from html.parser import HTMLParser
from io import BytesIO
from pathlib import Path
from typing import Any, Protocol
from urllib.parse import urljoin, urlsplit

from openpyxl import load_workbook
from PIL import Image, UnidentifiedImageError


MAX_IMAGE_BYTES = 10 * 1024 * 1024
MAX_IMAGE_PIXELS = 40_000_000
MAX_REDIRECTS = 3
_CONTROL = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f\x7f]")
_SHOP_IN_TEXT = re.compile(r"https://(?:www\.)?etsy\.com/shop/[A-Za-z0-9_-]+", re.IGNORECASE)
_LISTING_IN_HTML = re.compile(
    r"(?<![A-Za-z0-9._-])(?:https://(?:www\.)?etsy\.com)?/listing/[0-9]+(?:/[A-Za-z0-9._~-]+)?/?",
    re.IGNORECASE,
)
_ALLOWED_IMAGE_MIME = {
    "image/jpeg": "JPEG",
    "image/jpg": "JPEG",
    "image/png": "PNG",
    "image/webp": "WEBP",
}


class ImageEvidenceError(ValueError):
    pass


class HttpResponse(Protocol):
    status_code: int
    url: object
    headers: Any

    def iter_bytes(self, chunk_size: int = 65_536): ...


class HttpClient(Protocol):
    def get(self, url: str, *, follow_redirects: bool = False) -> HttpResponse: ...


@dataclass(frozen=True)
class ListingSnapshot:
    canonical_url: str
    listing_id: str
    title: str
    description: str
    tags: list[str]
    text_facts: dict[str, list[str]]
    source_timestamp: datetime

    def compact_text(self) -> str:
        parts = [self.title, self.description]
        for field in sorted(self.text_facts):
            values = ", ".join(self.text_facts[field])
            if values:
                parts.append(f"{field}: {values}")
        return " | ".join(part for part in parts if part)[:20_000]


@dataclass(frozen=True)
class ImageEvidence:
    source_url: str
    path: Path
    sha256: str
    width: int
    height: int
    media_type: str


def _normalized_text(value: object, *, maximum: int) -> str:
    if not isinstance(value, str):
        return ""
    cleaned = " ".join(unicodedata.normalize("NFKC", html_module.unescape(value)).split())
    if _CONTROL.search(cleaned):
        cleaned = _CONTROL.sub(" ", cleaned)
        cleaned = " ".join(cleaned.split())
    return cleaned[:maximum].strip()


def _strict_etsy_url(value: str, pattern: re.Pattern[str]) -> tuple[str, re.Match[str]]:
    parsed = urlsplit(value.strip())
    if (
        parsed.scheme != "https"
        or (parsed.hostname or "").casefold() not in {"etsy.com", "www.etsy.com"}
        or parsed.username
        or parsed.password
        or parsed.port not in {None, 443}
        or parsed.query
        or parsed.fragment
    ):
        raise ValueError("unsafe Etsy URL")
    match = pattern.fullmatch(parsed.path)
    if match is None:
        raise ValueError("unexpected Etsy URL path")
    return parsed.path, match


def normalize_shop_url(value: str) -> str:
    _, match = _strict_etsy_url(value, re.compile(r"/shop/([A-Za-z0-9_-]+)/?"))
    return f"https://www.etsy.com/shop/{match.group(1)}"


def normalize_listing_url(value: str) -> tuple[str, str]:
    _, match = _strict_etsy_url(
        value,
        re.compile(r"/listing/([0-9]+)(?:/[A-Za-z0-9._~-]*)?/?"),
    )
    listing_id = match.group(1)
    return f"https://www.etsy.com/listing/{listing_id}", listing_id


def extract_shop_urls(workbook_path: str | Path) -> list[str]:
    source = Path(workbook_path)
    workbook = load_workbook(source, read_only=False, data_only=True)
    result: list[str] = []
    seen: set[str] = set()
    try:
        for sheet in workbook.worksheets:
            for row in sheet.iter_rows():
                for cell in row:
                    candidates: list[str] = []
                    if isinstance(cell.value, str):
                        candidates.extend(match.group(0) for match in _SHOP_IN_TEXT.finditer(cell.value))
                    hyperlink = getattr(cell, "hyperlink", None)
                    if hyperlink is not None and isinstance(hyperlink.target, str):
                        candidates.extend(match.group(0) for match in _SHOP_IN_TEXT.finditer(hyperlink.target))
                    for candidate in candidates:
                        try:
                            canonical = normalize_shop_url(candidate)
                        except ValueError:
                            continue
                        key = canonical.casefold()
                        if key not in seen:
                            seen.add(key)
                            result.append(canonical)
    finally:
        workbook.close()
    return result


def extract_listing_urls(html: str) -> list[str]:
    result: list[str] = []
    seen: set[str] = set()
    for match in _LISTING_IN_HTML.finditer(html):
        value = match.group(0)
        absolute = value if value.casefold().startswith("https://") else urljoin("https://www.etsy.com", value)
        try:
            canonical, listing_id = normalize_listing_url(absolute)
        except ValueError:
            continue
        if listing_id not in seen:
            seen.add(listing_id)
            result.append(canonical)
    return result


class _MetadataParser(HTMLParser):
    def __init__(self) -> None:
        super().__init__(convert_charrefs=True)
        self.title_parts: list[str] = []
        self.meta: dict[str, str] = {}
        self.jsonld: list[str] = []
        self._in_title = False
        self._in_jsonld = False
        self._script_parts: list[str] = []

    def handle_starttag(self, tag: str, attrs: list[tuple[str, str | None]]) -> None:
        attributes = {key.casefold(): value or "" for key, value in attrs}
        lowered = tag.casefold()
        if lowered == "title":
            self._in_title = True
        elif lowered == "meta":
            key = (attributes.get("property") or attributes.get("name") or "").casefold()
            content = attributes.get("content", "")
            if key and content and key not in self.meta:
                self.meta[key] = content
        elif lowered == "script" and attributes.get("type", "").casefold() == "application/ld+json":
            self._in_jsonld = True
            self._script_parts = []

    def handle_endtag(self, tag: str) -> None:
        lowered = tag.casefold()
        if lowered == "title":
            self._in_title = False
        elif lowered == "script" and self._in_jsonld:
            self._in_jsonld = False
            value = "".join(self._script_parts).strip()
            if value:
                self.jsonld.append(value)
            self._script_parts = []

    def handle_data(self, data: str) -> None:
        if self._in_title:
            self.title_parts.append(data)
        if self._in_jsonld:
            self._script_parts.append(data)


def _metadata(html: str) -> tuple[_MetadataParser, list[dict[str, Any]]]:
    parser = _MetadataParser()
    parser.feed(html[:2_000_000])
    objects: list[dict[str, Any]] = []

    def collect(value: Any) -> None:
        if isinstance(value, dict):
            objects.append(value)
            graph = value.get("@graph")
            if isinstance(graph, list):
                for item in graph:
                    collect(item)
        elif isinstance(value, list):
            for item in value:
                collect(item)

    for raw in parser.jsonld:
        try:
            collect(json.loads(raw))
        except (json.JSONDecodeError, TypeError, RecursionError):
            continue
    return parser, objects


def _is_product(value: dict[str, Any]) -> bool:
    kind = value.get("@type")
    if isinstance(kind, str):
        return kind.casefold() == "product"
    if isinstance(kind, list):
        return any(isinstance(item, str) and item.casefold() == "product" for item in kind)
    return False


def extract_listing_snapshot(
    html: str,
    canonical_url: str,
    fetched_at: datetime,
) -> ListingSnapshot:
    canonical, listing_id = normalize_listing_url(canonical_url)
    parser, objects = _metadata(html)
    product = next((item for item in objects if _is_product(item)), {})
    title = _normalized_text(product.get("name"), maximum=500)
    if not title:
        title = _normalized_text("".join(parser.title_parts), maximum=500)
        title = re.sub(r"\s+-\s+Etsy\s*$", "", title, flags=re.IGNORECASE)
    description = _normalized_text(product.get("description"), maximum=10_000)
    if not description:
        description = _normalized_text(parser.meta.get("description"), maximum=10_000)

    text_facts: dict[str, list[str]] = {}
    materials = product.get("material")
    if isinstance(materials, str):
        materials = [materials]
    if isinstance(materials, list):
        cleaned = [_normalized_text(item, maximum=200) for item in materials]
        cleaned = [item for item in cleaned if item]
        if cleaned:
            text_facts["materials"] = list(dict.fromkeys(cleaned))[:20]
    category = _normalized_text(product.get("category"), maximum=200)
    if category:
        text_facts["product_family"] = [category]

    raw_keywords = product.get("keywords")
    if isinstance(raw_keywords, str):
        raw_tags = re.split(r"[,|]", raw_keywords)
    elif isinstance(raw_keywords, list):
        raw_tags = raw_keywords
    else:
        raw_tags = []
    tags = [_normalized_text(item, maximum=100) for item in raw_tags]
    tags = [item for item in tags if item][:50]
    if not title and not description:
        raise ValueError("listing snapshot is empty")
    return ListingSnapshot(
        canonical_url=canonical,
        listing_id=listing_id,
        title=title,
        description=description,
        tags=list(dict.fromkeys(tags)),
        text_facts=text_facts,
        source_timestamp=fetched_at,
    )


def _image_url(value: object, *, page_url: str) -> str | None:
    if not isinstance(value, str) or not value.strip():
        return None
    candidate = urljoin(page_url, html_module.unescape(value.strip()))
    parsed = urlsplit(candidate)
    if parsed.scheme != "https" or parsed.username or parsed.password or parsed.port not in {None, 443}:
        return None
    return candidate


def select_main_image_url(html: str, page_url: str) -> str:
    parser, objects = _metadata(html)
    for product in (item for item in objects if _is_product(item)):
        images = product.get("image")
        if isinstance(images, str):
            images = [images]
        if isinstance(images, dict):
            images = [images.get("url") or images.get("contentUrl")]
        if isinstance(images, list):
            for item in images:
                if isinstance(item, dict):
                    item = item.get("url") or item.get("contentUrl")
                candidate = _image_url(item, page_url=page_url)
                if candidate:
                    return candidate
    fallback = _image_url(parser.meta.get("og:image"), page_url=page_url)
    if fallback:
        return fallback
    raise ValueError("listing main image is missing")


def _approved_image_url(value: str) -> str:
    parsed = urlsplit(value)
    host = (parsed.hostname or "").casefold()
    if (
        parsed.scheme != "https"
        or parsed.username
        or parsed.password
        or parsed.port not in {None, 443}
        or not (host == "etsystatic.com" or host.endswith(".etsystatic.com"))
    ):
        raise ImageEvidenceError("image URL or redirect is not approved")
    return value


def _header(headers: Any, name: str) -> str:
    if hasattr(headers, "get"):
        value = headers.get(name)
        if value is None:
            value = headers.get(name.casefold())
        return str(value or "").strip()
    return ""


def download_main_image(
    url: str,
    destination_root: str | Path,
    client: HttpClient,
) -> ImageEvidence:
    current = _approved_image_url(url)
    response: HttpResponse | None = None
    for redirect_count in range(MAX_REDIRECTS + 1):
        response = client.get(current, follow_redirects=False)
        status = int(response.status_code)
        if status in {301, 302, 303, 307, 308}:
            if redirect_count >= MAX_REDIRECTS:
                raise ImageEvidenceError("image redirect limit exceeded")
            location = _header(response.headers, "location")
            if not location:
                raise ImageEvidenceError("image redirect is missing a location")
            current = _approved_image_url(urljoin(current, location))
            continue
        if status != 200:
            raise ImageEvidenceError("image download failed")
        response_url = str(response.url)
        if response_url:
            _approved_image_url(response_url)
            current = response_url
        break
    if response is None or int(response.status_code) != 200:
        raise ImageEvidenceError("image download failed")

    media_type = _header(response.headers, "content-type").split(";", 1)[0].casefold()
    expected_format = _ALLOWED_IMAGE_MIME.get(media_type)
    if expected_format is None:
        raise ImageEvidenceError("image content type is not supported")
    declared = _header(response.headers, "content-length")
    if declared:
        try:
            if int(declared) < 1 or int(declared) > MAX_IMAGE_BYTES:
                raise ImageEvidenceError("image byte limit exceeded")
        except ValueError as exc:
            raise ImageEvidenceError("image content length is invalid") from exc

    body = bytearray()
    for chunk in response.iter_bytes(chunk_size=65_536):
        if not isinstance(chunk, (bytes, bytearray)):
            raise ImageEvidenceError("image response is invalid")
        body.extend(chunk)
        if len(body) > MAX_IMAGE_BYTES:
            raise ImageEvidenceError("image byte limit exceeded")
    if not body:
        raise ImageEvidenceError("image response is empty")

    try:
        with Image.open(BytesIO(body)) as source:
            actual_format = (source.format or "").upper()
            if actual_format != expected_format:
                raise ImageEvidenceError("image content type does not match its signature")
            width, height = source.size
            if width < 1 or height < 1 or width * height > MAX_IMAGE_PIXELS:
                raise ImageEvidenceError("image pixel limit exceeded")
            source.load()
            if source.mode in {"RGBA", "LA"} or (source.mode == "P" and "transparency" in source.info):
                rgba = source.convert("RGBA")
                normalized = Image.new("RGB", rgba.size, "white")
                normalized.paste(rgba, mask=rgba.getchannel("A"))
            else:
                normalized = source.convert("RGB")
    except ImageEvidenceError:
        raise
    except (UnidentifiedImageError, OSError, ValueError, Image.DecompressionBombError) as exc:
        raise ImageEvidenceError("image could not be decoded safely") from exc

    encoded = BytesIO()
    normalized.save(encoded, format="JPEG", quality=95, subsampling=0, optimize=False)
    data = encoded.getvalue()
    digest = hashlib.sha256(data).hexdigest()
    root = Path(destination_root).resolve()
    root.mkdir(parents=True, exist_ok=True)
    destination = (root / f"{digest}.jpg").resolve()
    if not destination.is_relative_to(root):
        raise ImageEvidenceError("image destination is invalid")
    if not destination.exists():
        with destination.open("xb") as output:
            output.write(data)
    elif destination.read_bytes() != data:
        raise ImageEvidenceError("image evidence hash collision")
    return ImageEvidence(
        source_url=current,
        path=destination,
        sha256=digest,
        width=width,
        height=height,
        media_type="image/jpeg",
    )
