from __future__ import annotations

import sys
sys.dont_write_bytecode = True

import argparse
import hashlib
import json
import re
import unicodedata
import zipfile
from datetime import date, datetime, time
from decimal import Decimal
from pathlib import Path
from typing import Any
from xml.etree import ElementTree

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import get_column_letter
for _stream in (sys.stdout, sys.stderr):
    try:
        _stream.reconfigure(encoding="utf-8", errors="replace")
    except Exception:
        pass
del _stream


OUTPUT_HEADERS = (
    "head titles",
    "13 tags",
    "SPECIFICATION",
    "Category",
    "Instructions for buyers",
)
MAX_FILE_BYTES = 50 * 1024 * 1024
MAX_UNCOMPRESSED_BYTES = 250 * 1024 * 1024
MAX_ZIP_ENTRIES = 10_000
MAX_ROWS = 20_000
MAX_COLUMNS = 500
MAX_CANDIDATE_FIELDS = 100
MAX_CANDIDATE_HEADER_CHARS = 256
MAX_CANDIDATE_VALUE_CHARS = 8_000
MAX_ROW_CONTEXT_BYTES = 64 * 1024
_UNSUPPORTED_PART_PREFIXES = (
    "customxml/",
    "xl/activex/",
    "xl/controls/",
    "xl/ctrlprops/",
    "xl/customdata/",
    "xl/embeddings/",
    "xl/externallinks/",
    "xl/slicers/",
    "xl/slicercaches/",
    "xl/threadedcomments/",
    "xl/webextensions/",
)
_UNSUPPORTED_PARTS = {
    "xl/connections.xml",
    "xl/model/model.xml",
    "xl/vbaproject.bin",
}
_UNSUPPORTED_RELATIONSHIP_KINDS = {
    "activexcontrolbinary", "activexcontrol", "connections", "control", "ctrlprop", "customxml",
    "externallink", "externallinkpath", "oleobject", "package", "slicer", "slicercache",
    "threadedcomment", "vmlDrawing".casefold(), "webextension",
}
_REL_NS = "http://schemas.openxmlformats.org/package/2006/relationships"
_HYPERLINK_RELATIONSHIP = "http://schemas.openxmlformats.org/officeDocument/2006/relationships/hyperlink"
_INTERNAL_HEADER_TERMS = (
    "cost", "price cost", "profit", "margin", "revenue", "commission", "fee",
    "shipping cost", "logistics", "freight", "warehouse", "supplier", "finance",
    "internal", "采购", "成本", "利润", "毛利", "佣金", "费用", "物流", "运费",
    "仓库", "供应商", "财务", "内部", "拿货价", "销售价", "面价", "底价", "获利",
    "淘宝链接", "采购链接", "供应商链接", "来源链接",
    "competitor", "comparison source", "raw evidence", "竞品", "对标链接", "原始证据",
)
_INSTRUCTION_MARKERS = (
    "instruction", "please fill", "fill in", "do not edit", "template", "example only",
    "说明", "填写", "请填", "模板", "示例", "勿改",
)


class WorkbookError(RuntimeError):
    def __init__(self, code: str, message: str, *, details: dict[str, Any] | None = None) -> None:
        super().__init__(message)
        self.code = code
        self.message = message
        self.details = details or {}

    def as_dict(self) -> dict[str, Any]:
        return {"error": {"code": self.code, "message": self.message, "details": self.details}}


def normalize_header(value: object) -> str:
    """NFKC-normalize and collapse every Unicode whitespace run; matching remains case-sensitive."""
    if not isinstance(value, str):
        return ""
    return re.sub(r"\s+", " ", unicodedata.normalize("NFKC", value)).strip()


def _semantic_header(value: object) -> str:
    return normalize_header(value).casefold()


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for chunk in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _package_inventory(path: Path, *, validate_supported: bool) -> dict[str, tuple[Any, ...]]:
    relationships: list[tuple[str, str, str, str]] = []
    with zipfile.ZipFile(path) as archive:
        members = tuple(sorted(item.filename.replace("\\", "/") for item in archive.infolist()))
        for item in archive.infolist():
            normalized_name = item.filename.replace("\\", "/").casefold().lstrip("/")
            if validate_supported and (
                normalized_name in _UNSUPPORTED_PARTS
                or normalized_name.startswith(_UNSUPPORTED_PART_PREFIXES)
                or normalized_name.endswith(".vml")
            ):
                raise WorkbookError(
                    "unsupported_workbook_part",
                    "The workbook contains an advanced package part that cannot be preserved safely.",
                    details={"part": item.filename},
                )
            if not normalized_name.endswith(".rels"):
                continue
            try:
                root = ElementTree.fromstring(archive.read(item.filename))
            except ElementTree.ParseError as exc:
                raise WorkbookError("invalid_workbook", "The workbook contains invalid package relationships.") from exc
            for relationship in root.findall(f"{{{_REL_NS}}}Relationship"):
                relationship_type = str(relationship.attrib.get("Type", ""))
                target = str(relationship.attrib.get("Target", ""))
                target_mode = str(relationship.attrib.get("TargetMode", ""))
                if not relationship_type or not target:
                    raise WorkbookError("invalid_workbook", "A package relationship is missing its type or target.")
                relationships.append((item.filename.replace("\\", "/"), relationship_type, target, target_mode))
                relationship_kind = relationship_type.rsplit("/", 1)[-1].casefold()
                external_is_ordinary_hyperlink = target_mode.casefold() == "external" and relationship_type == _HYPERLINK_RELATIONSHIP
                if validate_supported and (
                    relationship_kind in _UNSUPPORTED_RELATIONSHIP_KINDS
                    or (target_mode and not external_is_ordinary_hyperlink)
                ):
                    raise WorkbookError(
                        "unsupported_workbook_part",
                        "The workbook contains an advanced or external relationship that cannot be preserved safely.",
                        details={"part": item.filename},
                    )
    return {"members": members, "relationships": tuple(sorted(relationships))}


def _validate_container(path: Path) -> None:
    if path.suffix.casefold() != ".xlsx":
        raise WorkbookError("unsupported_workbook_type", "Only .xlsx workbooks are supported.")
    if not path.is_file():
        raise WorkbookError("workbook_not_found", "The source workbook does not exist.")
    if path.stat().st_size > MAX_FILE_BYTES:
        raise WorkbookError("workbook_too_large", "The source workbook exceeds the size limit.")
    try:
        with zipfile.ZipFile(path) as archive:
            infos = archive.infolist()
            if len(infos) > MAX_ZIP_ENTRIES:
                raise WorkbookError("unsafe_workbook_archive", "The workbook contains too many archive entries.")
            total = sum(item.file_size for item in infos)
            if total > MAX_UNCOMPRESSED_BYTES:
                raise WorkbookError("unsafe_workbook_archive", "The workbook expands beyond the safe limit.")
            for item in infos:
                if item.file_size > 10 * 1024 * 1024 and item.compress_size and item.file_size / item.compress_size > 200:
                    raise WorkbookError("unsafe_workbook_archive", "The workbook contains a suspicious archive entry.")
            _package_inventory(path, validate_supported=True)
    except zipfile.BadZipFile as exc:
        raise WorkbookError("invalid_workbook", "The workbook is not a valid XLSX archive.") from exc


def _merged_range_for(ws, row: int, column: int):
    for merged in ws.merged_cells.ranges:
        if merged.min_row <= row <= merged.max_row and merged.min_col <= column <= merged.max_col:
            return merged
    return None


def _locate_header(wb) -> tuple[Any, int, dict[str, int]]:
    exact: list[tuple[Any, int, dict[str, int]]] = []
    diagnostic: list[tuple[int, int, Any, int, dict[str, list[int]]]] = []
    required = set(OUTPUT_HEADERS)
    for ws in wb.worksheets:
        if ws.max_row > MAX_ROWS or ws.max_column > MAX_COLUMNS:
            raise WorkbookError("worksheet_dimensions_exceeded", "A worksheet exceeds safe dimension limits.", details={"sheet": ws.title})
        for row_no in range(1, ws.max_row + 1):
            positions: dict[str, list[int]] = {header: [] for header in OUTPUT_HEADERS}
            for cell in ws[row_no]:
                normalized = normalize_header(cell.value)
                if normalized in required:
                    positions[normalized].append(cell.column)
            matched = sum(bool(value) for value in positions.values())
            duplicates = sum(max(0, len(value) - 1) for value in positions.values())
            if matched:
                diagnostic.append((matched, -duplicates, ws, row_no, positions))
            if matched == len(required) and duplicates == 0:
                exact.append((ws, row_no, {header: positions[header][0] for header in OUTPUT_HEADERS}))
    if len(exact) > 1:
        raise WorkbookError("ambiguous_header_location", "The output header row is ambiguous.", details={"locations": [{"sheet": ws.title, "row": row} for ws, row, _ in exact]})
    if len(exact) == 1:
        ws, row_no, mapping = exact[0]
        merged = {header: str(_merged_range_for(ws, row_no, column)) for header, column in mapping.items() if _merged_range_for(ws, row_no, column)}
        if merged:
            raise WorkbookError("merged_output_header", "An output header is part of a merged range.", details={"headers": merged})
        return ws, row_no, mapping
    if not diagnostic:
        raise WorkbookError("missing_output_header", "No row contains the required output headers.", details={"missing": list(OUTPUT_HEADERS)})
    _, _, ws, row_no, positions = sorted(diagnostic, key=lambda item: (item[0], item[1]), reverse=True)[0]
    duplicate = [header for header, columns in positions.items() if len(columns) > 1]
    if duplicate:
        raise WorkbookError("duplicate_output_header", "An output header occurs more than once in the candidate row.", details={"sheet": ws.title, "row": row_no, "headers": duplicate})
    merged = {}
    for header, columns in positions.items():
        if columns:
            merged_range = _merged_range_for(ws, row_no, columns[0])
            if merged_range:
                merged[header] = str(merged_range)
    if merged:
        raise WorkbookError("merged_output_header", "An output header is part of a merged range.", details={"headers": merged})
    missing = [header for header, columns in positions.items() if not columns]
    raise WorkbookError("missing_output_header", "The candidate header row is missing output headers.", details={"sheet": ws.title, "row": row_no, "missing": missing})


def _json_value(value: Any) -> Any:
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    if isinstance(value, (datetime, date, time)):
        return value.isoformat()
    if isinstance(value, Decimal):
        return str(value)
    return str(value)


def _cell_type(cell) -> str:
    return {"s": "text", "n": "number", "b": "boolean", "d": "date", "f": "formula", "e": "error"}.get(cell.data_type, "text")


def _is_internal(header: object) -> bool:
    semantic = _semantic_header(header)
    words = set(re.findall(r"[a-z0-9]+", semantic))
    for term in _INTERNAL_HEADER_TERMS:
        if re.search(r"[\u3400-\u9fff]", term):
            if term in semantic:
                return True
        else:
            term_words = term.split()
            if len(term_words) == 1 and term in words:
                return True
            if len(term_words) > 1 and re.search(rf"(?<![a-z0-9]){re.escape(term)}(?![a-z0-9])", semantic):
                return True
    return False


def _is_instruction_only(fields: list[dict[str, Any]]) -> bool:
    texts = [normalize_header(field.get("value")) for field in fields if field.get("value") not in (None, "")]
    if not texts:
        return True
    combined = " ".join(texts).casefold()
    first = texts[0].casefold()
    strong_phrase = any(marker in combined for marker in ("please fill", "fill in", "do not edit", "example only", "请填", "填写", "勿改"))
    explicit_label = first in {"instruction", "instructions", "说明", "模板", "示例"}
    return strong_phrase and explicit_label


def _extract_images(ws, product_rows: set[int], operation_dir: Path) -> dict[int, list[str]]:
    grouped = {row: [] for row in product_rows}
    images_dir = operation_dir / "images"
    for image in ws._images:
        anchor = getattr(image, "anchor", None)
        marker = getattr(anchor, "_from", None)
        if marker is None:
            continue
        row_no = marker.row + 1
        if row_no not in product_rows:
            continue
        grouped[row_no].append(image)
    if any(grouped.values()):
        images_dir.mkdir(parents=True, exist_ok=True)
    result = {row: [] for row in product_rows}
    for row_no in sorted(grouped):
        for index, image in enumerate(grouped[row_no], 1):
            extension = str(getattr(image, "format", "png") or "png").casefold()
            if extension not in {"png", "jpeg", "jpg", "gif", "bmp"}:
                extension = "bin"
            name = f"row-{row_no:06d}-image-{index:03d}.{extension}"
            destination = images_dir / name
            destination.write_bytes(image._data())
            result[row_no].append(str(destination.resolve()))
    return result


def inspect_workbook(source_path: str | Path, operation_dir: str | Path) -> dict[str, Any]:
    source = Path(source_path).resolve(strict=False)
    operation = Path(operation_dir).resolve()
    _validate_container(source)
    source_sha = _sha256(source)
    operation.mkdir(parents=True, exist_ok=True)
    try:
        # Formula text is inspected but never evaluated; external workbook links are not retained.
        workbook = load_workbook(source, data_only=False, keep_links=False, read_only=False)
    except Exception as exc:
        raise WorkbookError("invalid_workbook", "The workbook could not be opened safely.") from exc
    ws, header_row, output_mapping = _locate_header(workbook)
    output_columns = set(output_mapping.values())
    rows: list[dict[str, Any]] = []
    for row_no in range(header_row + 1, ws.max_row + 1):
        candidate_fields: list[dict[str, Any]] = []
        meaningful: list[Any] = []
        for column in range(1, ws.max_column + 1):
            if column in output_columns:
                continue
            header = ws.cell(header_row, column).value
            cell = ws.cell(row_no, column)
            if not normalize_header(header) or cell.value in (None, "") or isinstance(cell, MergedCell):
                continue
            meaningful.append(cell.value)
            if _is_internal(header):
                continue
            normalized_header = normalize_header(header)
            json_value = _json_value(cell.value)
            if len(normalized_header) > MAX_CANDIDATE_HEADER_CHARS:
                raise WorkbookError("workbook_input_limit_exceeded", "A candidate field header exceeds the safe input limit.", details={"row": row_no})
            if isinstance(json_value, str) and len(json_value) > MAX_CANDIDATE_VALUE_CHARS:
                raise WorkbookError("workbook_input_limit_exceeded", "A candidate field value exceeds the safe input limit.", details={"row": row_no, "header": normalized_header})
            candidate_fields.append({"header": normalized_header, "value": json_value, "type": _cell_type(cell)})
            if len(candidate_fields) > MAX_CANDIDATE_FIELDS:
                raise WorkbookError("workbook_input_limit_exceeded", "A product row contains too many candidate fields.", details={"row": row_no})
        if not meaningful or not candidate_fields or _is_instruction_only(candidate_fields):
            continue
        context = json.dumps(candidate_fields, ensure_ascii=False, sort_keys=True, separators=(",", ":"))
        if len(context.encode("utf-8")) > MAX_ROW_CONTEXT_BYTES:
            raise WorkbookError("workbook_input_limit_exceeded", "A product row exceeds the safe context limit.", details={"row": row_no})
        context_hash = hashlib.sha256(context.encode("utf-8")).hexdigest()
        row_id = hashlib.sha256(f"{source_sha}:{ws.title}:{row_no}:{context_hash}".encode("utf-8")).hexdigest()
        rows.append({
            "row_id": row_id,
            "row_number": row_no,
            "context_hash": context_hash,
            "context": context,
            "candidate_fields": candidate_fields,
            "image_paths": [],
            "visual_context": None,
            "warnings": [],
        })
    images = _extract_images(ws, {item["row_number"] for item in rows}, operation)
    for item in rows:
        item["image_paths"] = images[item["row_number"]]
        if len(item["image_paths"]) > 1:
            item["warnings"].append(f"{len(item['image_paths'])} images are associated with this row; only the first will be sent to the model.")
    if not rows:
        raise WorkbookError("no_product_rows", "No product rows were found below the selected output header row.")
    return {
        "manifest_version": 1,
        "source_path": str(source),
        "source_sha256": source_sha,
        "sheet": ws.title,
        "header_row": header_row,
        "output_columns": {header: get_column_letter(column) for header, column in output_mapping.items()},
        "rows": rows,
        "warnings": [],
    }


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("source")
    parser.add_argument("operation_dir")
    parser.add_argument("--manifest")
    args = parser.parse_args()
    try:
        manifest = inspect_workbook(args.source, args.operation_dir)
        text = json.dumps(manifest, ensure_ascii=False, indent=2)
        if args.manifest:
            Path(args.manifest).write_text(text, encoding="utf-8")
        print(text)
        return 0
    except WorkbookError as exc:
        print(json.dumps(exc.as_dict(), ensure_ascii=False))
        return 2


if __name__ == "__main__":
    raise SystemExit(main())
