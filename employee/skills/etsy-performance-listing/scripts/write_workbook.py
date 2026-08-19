from __future__ import annotations

import sys
sys.dont_write_bytecode = True

import argparse
import hashlib
import importlib.util
import json
import os
import posixpath
import re
import shutil
import tempfile
from io import BytesIO
from pathlib import Path
from typing import Any
from xml.etree import ElementTree
from zipfile import ZIP_DEFLATED, ZipFile

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter
for _stream in (sys.stdout, sys.stderr):
    try:
        _stream.reconfigure(encoding="utf-8", errors="replace")
    except Exception:
        pass
del _stream


HEADER_TO_KEY = {
    "head titles": "head_titles",
    "13 tags": "tags",
    "SPECIFICATION": "specification",
    "Category": "category",
    "Instructions for buyers": "instructions_for_buyers",
}

_SPREADSHEET_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
_DOCUMENT_REL_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
_PACKAGE_REL_NS = "http://schemas.openxmlformats.org/package/2006/relationships"
_XML_NS = "http://www.w3.org/XML/1998/namespace"
_CELL_REFERENCE = re.compile(r"^([A-Z]+)([1-9][0-9]*)$")


class WorkbookWriteError(RuntimeError):
    def __init__(self, code: str, message: str, *, details: dict[str, Any] | None = None) -> None:
        super().__init__(message)
        self.code = code
        self.details = details or {}

    def as_dict(self) -> dict[str, Any]:
        return {"error": {"code": self.code, "message": str(self), "details": self.details}}


def _load_sibling(name: str):
    path = Path(__file__).with_name(f"{name}.py")
    spec = importlib.util.spec_from_file_location(f"etsy_excel_{name}", path)
    if spec is None or spec.loader is None:
        raise RuntimeError(f"Could not load {path.name}")
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for chunk in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _safe_excel_string(value: Any) -> str:
    if isinstance(value, list):
        value = ", ".join(str(item) for item in value)
    text = str(value)
    if text.startswith(("=", "+", "-", "@")):
        raise WorkbookWriteError("unsafe_output_value", "An output value begins with an Excel formula prefix.")
    return text


def _validate_manifest_mapping(workbook, manifest: dict[str, Any]) -> tuple[Any, dict[str, int]]:
    inspect = _load_sibling("inspect_workbook")
    sheet = manifest.get("sheet")
    if not isinstance(sheet, str) or sheet not in workbook.sheetnames:
        raise WorkbookWriteError("manifest_mismatch", "The manifest sheet is not present in the copied workbook.")
    ws = workbook[sheet]
    header_row = manifest.get("header_row")
    mapping = manifest.get("output_columns")
    if not isinstance(header_row, int) or not isinstance(mapping, dict) or set(mapping) != set(HEADER_TO_KEY):
        raise WorkbookWriteError("manifest_mismatch", "The manifest output mapping is invalid.")
    columns: dict[str, int] = {}
    for header, letter in mapping.items():
        try:
            column = column_index_from_string(letter)
        except (TypeError, ValueError) as exc:
            raise WorkbookWriteError("manifest_mismatch", "The manifest contains an invalid output column.") from exc
        if inspect.normalize_header(ws.cell(header_row, column).value) != header:
            raise WorkbookWriteError("manifest_mismatch", "The copied workbook no longer matches the manifest output mapping.")
        if inspect._merged_range_for(ws, header_row, column):
            raise WorkbookWriteError("manifest_mismatch", "An output header became merged.")
        columns[header] = column
    return ws, columns


def _worksheet_part(archive: ZipFile, sheet_name: str) -> str:
    workbook_root = ElementTree.fromstring(archive.read("xl/workbook.xml"))
    relationship_id: str | None = None
    for sheet in workbook_root.findall(f".//{{{_SPREADSHEET_NS}}}sheet"):
        if sheet.get("name") == sheet_name:
            relationship_id = sheet.get(f"{{{_DOCUMENT_REL_NS}}}id")
            break
    if not relationship_id:
        raise WorkbookWriteError("manifest_mismatch", "The manifest sheet relationship is missing.")
    relationships = ElementTree.fromstring(archive.read("xl/_rels/workbook.xml.rels"))
    target: str | None = None
    for relationship in relationships.findall(f"{{{_PACKAGE_REL_NS}}}Relationship"):
        if relationship.get("Id") == relationship_id:
            if relationship.get("TargetMode") == "External":
                raise WorkbookWriteError("manifest_mismatch", "The manifest sheet relationship is external.")
            target = relationship.get("Target")
            break
    if not target:
        raise WorkbookWriteError("manifest_mismatch", "The manifest sheet target is missing.")
    part = posixpath.normpath(target.lstrip("/") if target.startswith("/") else posixpath.join("xl", target))
    if not part.startswith("xl/worksheets/") or part not in archive.namelist():
        raise WorkbookWriteError("manifest_mismatch", "The manifest sheet target is unsafe or missing.")
    return part


def _registered_namespaces(xml: bytes) -> None:
    for _event, (prefix, uri) in ElementTree.iterparse(BytesIO(xml), events=("start-ns",)):
        try:
            ElementTree.register_namespace(prefix or "", uri)
        except ValueError as exc:
            raise WorkbookWriteError("workbook_write_failed", "The worksheet namespace map is invalid.") from exc


def _cell_column(cell: ElementTree.Element) -> int:
    match = _CELL_REFERENCE.fullmatch(cell.get("r", ""))
    return column_index_from_string(match.group(1)) if match else 1_000_000


def _patch_worksheet_xml(xml: bytes, values: dict[str, str]) -> bytes:
    _registered_namespaces(xml)
    root = ElementTree.fromstring(xml)
    sheet_data = root.find(f"{{{_SPREADSHEET_NS}}}sheetData")
    if sheet_data is None:
        raise WorkbookWriteError("manifest_mismatch", "The manifest worksheet has no sheet data.")
    rows = {int(row.get("r")): row for row in sheet_data.findall(f"{{{_SPREADSHEET_NS}}}row") if (row.get("r") or "").isdigit()}
    for address, value in sorted(
        values.items(),
        key=lambda item: (
            int(_CELL_REFERENCE.fullmatch(item[0]).group(2)),
            _cell_column(ElementTree.Element("c", {"r": item[0]})),
        ),
    ):
        match = _CELL_REFERENCE.fullmatch(address)
        if match is None:
            raise WorkbookWriteError("manifest_mismatch", "An output cell address is invalid.")
        row_number = int(match.group(2))
        row = rows.get(row_number)
        if row is None:
            row = ElementTree.Element(f"{{{_SPREADSHEET_NS}}}row", {"r": str(row_number)})
            insertion = next((index for index, current in enumerate(sheet_data) if int(current.get("r", "0")) > row_number), len(sheet_data))
            sheet_data.insert(insertion, row)
            rows[row_number] = row
        cells = row.findall(f"{{{_SPREADSHEET_NS}}}c")
        cell = next((item for item in cells if item.get("r") == address), None)
        if cell is None:
            cell = ElementTree.Element(f"{{{_SPREADSHEET_NS}}}c", {"r": address})
            column = column_index_from_string(match.group(1))
            insertion = next((index for index, current in enumerate(row) if current.tag == f"{{{_SPREADSHEET_NS}}}c" and _cell_column(current) > column), len(row))
            row.insert(insertion, cell)
        for child in list(cell):
            cell.remove(child)
        cell.set("t", "inlineStr")
        inline = ElementTree.SubElement(cell, f"{{{_SPREADSHEET_NS}}}is")
        text = ElementTree.SubElement(inline, f"{{{_SPREADSHEET_NS}}}t")
        if value != value.strip():
            text.set(f"{{{_XML_NS}}}space", "preserve")
        text.text = value
    return ElementTree.tostring(root, encoding="utf-8", xml_declaration=True)


def _rewrite_workbook_package(path: Path, sheet_name: str, values: dict[str, str]) -> None:
    fd, replacement_name = tempfile.mkstemp(prefix=".listing-package-", suffix=".xlsx", dir=path.parent)
    os.close(fd)
    replacement = Path(replacement_name)
    try:
        with ZipFile(path, "r") as source_archive, ZipFile(replacement, "w", compression=ZIP_DEFLATED) as output_archive:
            worksheet_part = _worksheet_part(source_archive, sheet_name)
            output_archive.comment = source_archive.comment
            for member in source_archive.infolist():
                payload = source_archive.read(member.filename)
                if member.filename == worksheet_part:
                    payload = _patch_worksheet_xml(payload, values)
                output_archive.writestr(member, payload)
        os.replace(replacement, path)
    finally:
        replacement.unlink(missing_ok=True)


def write_workbook(
    source_path: str | Path,
    output_dir: str | Path,
    manifest: dict[str, Any],
    row_results: dict[str, dict[str, Any]],
    *,
    rules: dict[str, Any],
    expected_rule_version: str,
) -> dict[str, Any]:
    source = Path(source_path).resolve(strict=True)
    destination_dir = Path(output_dir).resolve()
    if source.suffix.casefold() != ".xlsx":
        raise WorkbookWriteError("unsupported_workbook_type", "Only .xlsx workbooks are supported.")
    inspect = _load_sibling("inspect_workbook")
    try:
        inspect._validate_container(source)
    except inspect.WorkbookError as exc:
        raise WorkbookWriteError(exc.code, str(exc), details=exc.details) from exc
    source_package_inventory = inspect._package_preservation_signature(source, validate_supported=True)
    if _sha256(source) != manifest.get("source_sha256"):
        raise WorkbookWriteError("source_hash_mismatch", "The source workbook changed after inspection.")
    if not isinstance(expected_rule_version, str) or not expected_rule_version.strip():
        raise WorkbookWriteError("invalid_rules", "An expected rule version is required.")
    active_rules = dict(rules)
    configured_version = active_rules.get("rule_version")
    if configured_version is not None and configured_version != expected_rule_version:
        raise WorkbookWriteError("invalid_rules", "The configured and expected rule versions do not match.")
    active_rules["rule_version"] = expected_rule_version
    destination_dir.mkdir(parents=True, exist_ok=True)
    final = destination_dir / f"{source.stem}-generated-{manifest['source_sha256'][:12]}.xlsx"
    if final.exists():
        raise WorkbookWriteError("output_exists", "The generated batch workbook already exists.")
    rows = manifest.get("rows")
    if not isinstance(rows, list):
        raise WorkbookWriteError("manifest_mismatch", "The manifest row list is invalid.")
    row_ids = [item.get("row_id") for item in rows if isinstance(item, dict)]
    row_numbers = [item.get("row_number") for item in rows if isinstance(item, dict)]
    if len(row_ids) != len(rows) or len(set(row_ids)) != len(row_ids) or len(set(row_numbers)) != len(row_numbers):
        raise WorkbookWriteError("manifest_mismatch", "The manifest contains duplicate or invalid row identities.")
    manifest_rows = {item["row_id"]: item for item in rows}
    if not set(row_results).issubset(manifest_rows):
        raise WorkbookWriteError("unknown_row", "A generated result does not belong to this manifest.")
    validator = _load_sibling("validate_output")
    validated_results: dict[str, dict[str, Any]] = {}
    for row_id, result in row_results.items():
        try:
            validated_results[row_id] = validator.validate_generated(result, active_rules)
        except validator.OutputValidationError as exc:
            raise WorkbookWriteError("invalid_result", "A generated result failed strict output validation.", details=exc.as_dict()["error"]["details"]) from exc
    temp_path: Path | None = None
    try:
        fd, temp_name = tempfile.mkstemp(prefix=".listing-", suffix=".xlsx", dir=destination_dir)
        os.close(fd)
        temp_path = Path(temp_name)
        shutil.copyfile(source, temp_path)
        if _sha256(temp_path) != manifest.get("source_sha256"):
            raise WorkbookWriteError("copy_hash_mismatch", "The temporary workbook copy does not match the inspected source.")
        with tempfile.TemporaryDirectory(prefix="listing-reinspect-", dir=destination_dir) as reinspection_dir:
            derived = inspect.inspect_workbook(temp_path, reinspection_dir)
        identity_keys = ("row_id", "row_number", "context_hash", "context", "candidate_fields")
        if derived.get("sheet") != manifest.get("sheet") or derived.get("header_row") != manifest.get("header_row") or derived.get("output_columns") != manifest.get("output_columns"):
            raise WorkbookWriteError("manifest_mismatch", "The manifest header binding does not match the copied workbook.")
        derived_rows = derived.get("rows", [])
        if len(derived_rows) != len(rows):
            raise WorkbookWriteError("manifest_mismatch", "The manifest row set does not match the copied workbook.")
        for expected, actual in zip(rows, derived_rows, strict=True):
            if any(expected.get(key) != actual.get(key) for key in identity_keys):
                raise WorkbookWriteError("manifest_mismatch", "A manifest row identity or context does not match the copied workbook.")
        workbook = load_workbook(temp_path, data_only=False, keep_links=True, read_only=False)
        ws, columns = _validate_manifest_mapping(workbook, manifest)
        changed: list[str] = []
        expected_cells: dict[str, str] = {}
        for row_id in sorted(validated_results):
            row_number = manifest_rows[row_id].get("row_number")
            if not isinstance(row_number, int) or row_number <= manifest["header_row"]:
                raise WorkbookWriteError("manifest_mismatch", "The manifest contains an invalid product row.")
            result = validated_results[row_id]
            for header, key in HEADER_TO_KEY.items():
                if key not in result:
                    raise WorkbookWriteError("invalid_result", f"The generated result is missing {key}.")
                value = _safe_excel_string(result[key])
                address = f"{ws.title}!{get_column_letter(columns[header])}{row_number}"
                changed.append(address)
                expected_cells[address] = value
        workbook.close()
        package_values = {address.split("!", 1)[1]: value for address, value in expected_cells.items()}
        _rewrite_workbook_package(temp_path, manifest["sheet"], package_values)
        output_package_inventory = inspect._package_preservation_signature(temp_path, validate_supported=True)
        if output_package_inventory != source_package_inventory:
            raise WorkbookWriteError(
                "package_preservation_failed",
                "The generated workbook package or relationship inventory changed unexpectedly.",
            )
        reopened = load_workbook(temp_path, data_only=False, keep_links=True, read_only=False)
        reopened_ws, reopened_columns = _validate_manifest_mapping(reopened, manifest)
        for row_id, result in validated_results.items():
            row_number = manifest_rows[row_id]["row_number"]
            for header in HEADER_TO_KEY:
                address = f"{reopened_ws.title}!{get_column_letter(reopened_columns[header])}{row_number}"
                if reopened_ws.cell(row_number, reopened_columns[header]).value != expected_cells[address]:
                    raise WorkbookWriteError("output_verification_failed", "A generated cell did not retain its exact validated value.", details={"cell": address})
        if _sha256(source) != manifest.get("source_sha256"):
            raise WorkbookWriteError("source_hash_mismatch", "The source workbook changed during writing.")
        os.replace(temp_path, final)
        temp_path = None
        return {"output_path": str(final), "output_sha256": _sha256(final), "changed_cells": sorted(changed)}
    except WorkbookWriteError:
        raise
    except Exception as exc:
        raise WorkbookWriteError("workbook_write_failed", "The output workbook could not be written safely.") from exc
    finally:
        if temp_path is not None:
            try:
                temp_path.unlink(missing_ok=True)
            except OSError:
                # Cleanup is best-effort and must never replace the primary write error.
                pass


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("source")
    parser.add_argument("output_dir")
    parser.add_argument("manifest")
    parser.add_argument("results")
    parser.add_argument("--rules", required=True)
    parser.add_argument("--expected-rule-version", required=True)
    args = parser.parse_args()
    try:
        report = write_workbook(
            args.source,
            args.output_dir,
            json.loads(Path(args.manifest).read_text(encoding="utf-8")),
            json.loads(Path(args.results).read_text(encoding="utf-8")),
            rules=json.loads(Path(args.rules).read_text(encoding="utf-8")),
            expected_rule_version=args.expected_rule_version,
        )
        print(json.dumps(report, ensure_ascii=False))
        return 0
    except (WorkbookWriteError, OSError, UnicodeError, json.JSONDecodeError) as exc:
        error = exc.as_dict() if isinstance(exc, WorkbookWriteError) else WorkbookWriteError("invalid_input", "The writer input could not be parsed.").as_dict()
        print(json.dumps(error, ensure_ascii=False))
        return 2


if __name__ == "__main__":
    raise SystemExit(main())
